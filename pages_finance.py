"""Finance pages — simplified for older colleagues."""

import streamlit as st
import pandas as pd
from datetime import datetime
import os, io, json, tempfile

from utils.database import (
    get_projects, get_clients, get_client_by_id, get_pending_approvals,
    approve_project, reject_project
)
from utils.receipt_pdf import generate_receipt_pdf
from utils.generate import generate_cash_receipt

STAGE_MAP = {'draft': '草稿', 'pending': '待审核', 'approved': '已开发票', 'rejected': '已驳回'}
CLOSURE_MAP = {'active': '进行中', 'pending_payment': '待收款', 'closed': '已结案'}
# 成本细项固定顺序：拍摄 → 餐饮交通 → 兼职执行 → 发布 → 补发
COST_ORDER = {'拍摄': 1, '餐饮交通': 2, '兼职执行': 3, '发布': 4, '补发': 5}


def _fmt_exec(val):
    """Convert execution period to Chinese format."""
    import re
    m = {'Jan':'01','Feb':'02','Mar':'03','Apr':'04','May':'05','Jun':'06',
         'Jul':'07','Aug':'08','Sep':'09','Oct':'10','Nov':'11','Dec':'12','January':'01','February':'02','March':'03','April':'04','June':'06','July':'07','August':'08','September':'09','October':'10','November':'11','December':'12'}
    for eng, num in m.items():
        if eng in val:
            val = val.replace(eng, num)
    val = re.sub(r'(\d{4})-(\d{1,2})\s*(到|-|–|~)\s*(\d{4})-(\d{1,2})', r'\1-\2 -\4-\5', val)
    # If only "YYYY-MM - YYYY-MM" format
    val = re.sub(r'(\d{4})-(\d{1,2})\s*(到|-|–)\s*(\d{2})$', r'\1-\2 -\1-\4', val) if '202' in val else val
    return val


def _fmt_date_val(val):
    """Consistent YYYY-MM-DD format."""
    if val is None: return ''
    if hasattr(val, 'strftime'): return val.strftime('%Y-%m-%d')
    return str(val)[:10]


def _fmt_cost_line(cost_json: str) -> str:
    if not cost_json: return ""
    try:
        items = json.loads(cost_json)
        items = sorted(items, key=lambda x: COST_ORDER.get(x.get('name',''), 99))
        return "、".join(f"{i['name']}({i.get('currency','RMB')}{i.get('amount',0):,.0f})" for i in items)
    except: return cost_json


def _inject_large_font_css():
    st.markdown("""
    <style>
    html, body, [class*="css"] { font-size: 16px !important; }
    div[data-testid="stMetricValue"] { font-size: 28px !important; }
    h1 { font-size: 28px !important; }
    h2 { font-size: 22px !important; }
    h3 { font-size: 18px !important; }
    button { font-size: 16px !important; padding: 10px 14px !important; }
    input, select { font-size: 16px !important; }
    table { font-size: 14px !important; }
    </style>
    """, unsafe_allow_html=True)


def page_overview():
    _inject_large_font_css()
    st.title("📊 项目总览")
    projects = get_projects(limit=500)
    if not projects: st.info("暂无项目"); return

    # Year-month filter
    months = sorted(set(
        f"20{p.get('project_code','')[4:6]}-{p.get('project_code','')[6:8]}"
        for p in projects if len(p.get('project_code','')) >= 8
    ), reverse=True)
    months = ['全部'] + months
    sel_month = st.selectbox("筛选年月", months)
    if sel_month != '全部':
        projects = [p for p in projects
                    if f"20{p.get('project_code','')[4:6]}-{p.get('project_code','')[6:8]}" == sel_month
                    and len(p.get('project_code','')) >= 8]

    # Summary cards
    pending_count = sum(1 for p in projects if p.get('status') == 'pending')
    approved_count = sum(1 for p in projects if p.get('status') == 'approved')
    received_full = sum(1 for p in projects if p.get('payment_received'))
    received_partial = sum(1 for p in projects if not p.get('payment_received') and (p.get('received_amount', 0) or 0) > 0)
    closed_count = sum(1 for p in projects if p.get('closure_status') == 'closed')
    total_cost = sum(p.get('estimated_cost',0) or 0 for p in projects)
    total_revenue = sum(p.get('amount',0) or 0 for p in projects)
    total_received = sum((p.get('received_amount', 0) or 0) for p in projects)
    need_receipt = sum(1 for p in projects if (p.get('received_amount', 0) or 0) > 0 and p.get('status')=='approved')
    c1,c2,c3,c4,c5,c6,c7,c8 = st.columns(8)
    c1.metric("⏳ 待审核", pending_count)
    c2.metric("✅ 已开发票", approved_count)
    c3.metric("💰 已到账", received_full)
    c4.metric("💵 部分到账", received_partial)
    c5.metric("🧾 待开收据", need_receipt)
    c6.metric("🔒 已结案", closed_count)
    c7.metric("💸 总成本", f"¥{total_cost:,.0f}")
    c8.metric("📈 总收入", f"${total_revenue:,.0f}")

    # Quick action: projects needing receipt (any received amount > 0)
    need_rec_projects = [p for p in projects if (p.get('received_amount', 0) or 0) > 0 and p.get('status')=='approved']
    if need_rec_projects:
        st.divider()
        st.subheader("🧾 待开收据项目")
        for p in need_rec_projects[:5]:
            c1,c2 = st.columns([4,1])
            with c1:
                rcvd = p.get('received_amount', 0) or 0
                total = p.get('amount', 0)
                pct = f"({rcvd/total*100:.0f}%)" if total > 0 else ""
                st.write(f"**{p.get('brand_name','')}** — {p.get('project_code','')} | 已到账 {p.get('currency','USD')} {rcvd:,.0f} / {total:,.0f} {pct}")
            with c2:
                if st.button("🧾 开收据", key=f"rec_{p['id']}", use_container_width=True):
                    st.session_state['receipt_project_id'] = p['id']
                    st.session_state.page = "receipt"
                    st.rerun()
    st.divider()

    # === Excel download ===
    if st.button("📥 下载成本明细表", use_container_width=True):
        _export_excel(projects)

    # === Table with merged cells ===
    st.subheader("项目明细")
    _render_table(projects)


def _render_table(projects):
    # 成本明细横向铺成列：每个成本项一列，成本项多自动多列（配合左右滑动）
    cost_cols = []
    seen = set()
    all_items = []
    for p in projects:
        try:
            items = json.loads(p.get('cost_breakdown', '') or '[]')
        except Exception:
            items = []
        all_items.extend(items)
    for it in sorted(all_items, key=lambda x: COST_ORDER.get(x.get('name', ''), 99)):
        nm = it.get('name', '')
        if nm and nm not in seen:
            seen.add(nm)
            cost_cols.append(nm)

    rows = []
    for seq_no, p in enumerate(projects, start=1):
        rcvd = p.get('received_amount', 0) or 0
        total = p.get('amount', 0) or 0
        remaining = max(total - rcvd, 0)
        if p.get('payment_received'):
            paid = '✅ 全款'
        elif rcvd > 0:
            paid = f'💵 {rcvd/total*100:.0f}%' if total else ''
        else:
            paid = ''
        try:
            items = json.loads(p.get('cost_breakdown', '') or '[]')
        except Exception:
            items = []
        cost_map = {it.get('name', ''): it.get('amount', 0) for it in items}

        row = {
            '序号': seq_no,
            '客户': p.get('client_short', '') or '',
            '项目名称': (p.get('project_name', '') or '')[:35],
            '编号': p.get('project_code', '') or '待分配',
            '金额': total,
            '已到账': rcvd,
            '待收': remaining,
            '本次到账': 0.0,
            '到账状态': paid,
        }
        for cc in cost_cols:
            row[cc] = cost_map.get(cc)
        row['总成本'] = p.get('estimated_cost', 0) or 0
        row['立项'] = '是' if p.get('feishu_approved') else '否'
        row['结案'] = CLOSURE_MAP.get(p.get('closure_status', 'active') or 'active', '')
        row['阶段'] = STAGE_MAP.get(p.get('status', ''), p.get('status', '?'))
        row['__id'] = p.get('id')
        rows.append(row)

    df = pd.DataFrame(rows)
    display_cols = [c for c in df.columns if c != '__id']
    read_only = [c for c in df.columns if c not in ('本次到账', '__id')]

    col_cfg = {
        '序号': st.column_config.TextColumn('序号', pinned=True, width='small'),
        '客户': st.column_config.TextColumn('客户', pinned=True, width='small'),
        '项目名称': st.column_config.TextColumn('项目名称', pinned=True, width='medium'),
        '编号': st.column_config.TextColumn('编号', width='medium'),
        '金额': st.column_config.NumberColumn('金额', format='%.0f', width='small'),
        '已到账': st.column_config.NumberColumn('已到账', format='%.0f', width='small'),
        '待收': st.column_config.NumberColumn('待收', format='%.0f', width='small'),
        '本次到账': st.column_config.NumberColumn('本次到账', format='%.0f', min_value=0.0, width='small'),
        '到账状态': st.column_config.TextColumn('到账状态', width='small'),
        '总成本': st.column_config.NumberColumn('总成本', format='%.0f', width='small'),
        '立项': st.column_config.TextColumn('立项', width='small'),
        '结案': st.column_config.TextColumn('结案', width='small'),
        '阶段': st.column_config.TextColumn('阶段', width='small'),
        '__id': None,
    }
    for cc in cost_cols:
        col_cfg[cc] = st.column_config.NumberColumn(cc, format='%.0f', width='small')

    st.caption('💡 在「本次到账」列输入金额 → 点「保存到账」批量入账。成本项多会自动多列，可左右滑动查看。')
    edited = st.data_editor(
        df,
        key='overview_editor',
        column_config=col_cfg,
        column_order=display_cols,
        disabled=read_only,
        hide_index=True,
        use_container_width=True,
        height=min(38 * (len(df) + 1) + 3, 600),
    )

    if st.button('💾 保存到账', type='primary', use_container_width=True):
        updates = []
        for _, row in edited.iterrows():
            amt = row.get('本次到账')
            if amt is None or pd.isna(amt):
                amt = 0.0
            amt = float(amt)
            pid = row.get('__id')
            if amt > 0 and pid is not None and not pd.isna(pid):
                updates.append((int(pid), amt))
        if not updates:
            st.warning('请先在「本次到账」列输入要入账的金额')
        else:
            from utils.database import get_connection
            n_ok = 0
            for pid, amt in updates:
                orig = next((x for x in projects if x.get('id') == pid), None)
                if orig is None:
                    continue
                new_total = (orig.get('received_amount', 0) or 0) + amt
                get_connection().table('projects').update({
                    'received_amount': new_total,
                    'payment_received': new_total >= (orig.get('amount', 0) or 0),
                    'received_date': datetime.now().strftime('%Y-%m-%d'),
                }).eq('id', pid).execute()
                n_ok += 1
            st.success(f'✅ 已入账 {n_ok} 笔。到账项目会出现在上方「待开收据」，去那里开收据即可。')
            st.rerun()


def _export_excel(projects):
    import openpyxl as xl
    from openpyxl.styles import Font, Alignment, Border, Side
    wb = xl.Workbook(); ws = wb.active; ws.title = "成本明细"
    hs = ['序号','客户','项目名称','编号','金额','执行周期','预计付款','成本细项','成本金额','总成本','立项','到账','结案','阶段']
    thin = Side(style='thin')
    for c,h in enumerate(hs,1):
        cell=ws.cell(1,c,h); cell.font=Font(bold=True); cell.alignment=Alignment(horizontal='center',vertical='center')
        cell.border=Border(bottom=thin)
    row=2
    for p in projects:
        stage=STAGE_MAP.get(p.get('status',''),'')
        closure=CLOSURE_MAP.get(p.get('closure_status',''),'')
        rcvd_amt = p.get('received_amount', 0) or 0
        if p.get('payment_received'): paid = '是'
        elif rcvd_amt > 0: paid = f'{rcvd_amt/p.get("amount", 1)*100:.0f}%'
        else: paid = '否'
        try: items=json.loads(p.get('cost_breakdown','') or '[]')
        except: items=[]
        if items:
            items = sorted(items, key=lambda x: COST_ORDER.get(x.get('name',''), 99))
        start_row = row
        total_cost = p.get('estimated_cost',0) or 0
        feishu = '是' if p.get('feishu_approved') else '否'
        exec_p = p.get('execution_period','') or ''
        exp_p = str(p.get('expected_payment_date','') or '')[:10]
        if items:
            for it in items:
                ws.cell(row,8,it.get('name','')); ws.cell(row,9,it.get('amount',0))
                row+=1
        else:
            ws.cell(row,9,total_cost)
            row+=1
        end_row = row - 1

        # Write merged project info
        ws.cell(start_row,1,p.get('id',''))
        ws.cell(start_row,2,p.get('client_short',''))
        ws.cell(start_row,3,(p.get('project_name','') or '')[:35])
        ws.cell(start_row,4,p.get('project_code',''))
        ws.cell(start_row,5,f"{p.get('currency','USD')} {p.get('amount',0):,.0f}")
        ws.cell(start_row,6,exec_p)
        ws.cell(start_row,7,exp_p)
        ws.cell(start_row,10,total_cost)
        ws.cell(start_row,11,feishu)
        ws.cell(start_row,12,paid)
        ws.cell(start_row,13,closure)
        ws.cell(start_row,14,stage)

        # Center
        for c in range(1,15):
            ws.cell(start_row,c).alignment=Alignment(horizontal='center',vertical='center')

        # Merge cells for multi-row projects
        if end_row > start_row:
            for c in [1,2,3,4,5,6,7,10,11,12,13,14]:
                ws.merge_cells(start_row=start_row, start_column=c, end_row=end_row, end_column=c)

        # Borders
        for r in range(start_row, end_row+1):
            for c in range(1,11):
                ws.cell(r,c).border=Border(bottom=Side(style='hair'))

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    st.download_button("📥 下载 Excel", buf, file_name="项目成本明细.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def page_approval():
    _inject_large_font_css()
    st.title("📋 审核状态")
    pending = get_pending_approvals()
    user=st.session_state.user

    if pending:
        st.subheader(f"共 {len(pending)} 个项目等你审核")
        for p in pending:
            with st.container(border=True):
                st.markdown(f"### {p.get('brand_name','')} — {p.get('project_name','')}")
                feishu_badge = "✅ 飞书已立项" if p.get('feishu_approved') else "⚠️ 未确认飞书立项"
                col_info,col_btn=st.columns([3,2])
                with col_info:
                    st.write(f"**{p.get('client_short','')}** | {p.get('currency','USD')} **{p.get('amount',0):,.2f}** | {feishu_badge}")
                    if p.get('estimated_cost'):
                        cd=_fmt_cost_line(p.get('cost_breakdown','') or '')
                        st.caption(f"预估成本: {p.get('estimated_cost',0):,.0f}"+(f"（{cd}）" if cd else ""))
                    st.caption(f"提交: {(p.get('created_at','') or '')[:10]}")
                with col_btn:
                    # Show stamped confirmation for download
                    if p.get('stamped_confirmation'):
                        with st.expander("📎 盖章确认函"):
                            import base64
                            st.download_button("📥 下载盖章确认函",
                                              base64.b64decode(p['stamped_confirmation']),
                                              file_name=f"{p.get('brand_name','')}-盖章确认函.pdf",
                                              mime="application/pdf")
                            try: st.image(base64.b64decode(p['stamped_confirmation']))
                            except: pass
                    with st.expander("📄 预览Invoice", expanded=True):
                        _show_invoice_preview(p)
                    _gen_invoice_dl(p)
                    if st.button("✅ 通过", key=f"ok_{p['id']}", use_container_width=True, type="primary"):
                        with st.spinner("生成盖章PDF..."):
                            try:
                                _regen_and_approve(p, user['id'])
                                st.success("已通过！")
                                code = p.get('project_code','')
                                month_str = code[8:10] if len(code)>=8 else ''
                                MONTHS = {'01':'Jan','02':'Feb','03':'Mar','04':'Apr','05':'May','06':'Jun',
                                          '07':'Jul','08':'Aug','09':'Sep','10':'Oct','11':'Nov','12':'Dec'}
                                m = MONTHS.get(month_str,'')
                                subj = f"Invoice for {p.get('brand_name','')} {m} Campaign / {p.get('project_code','')}"
                                body = (f"Dear all,\n\n"
                                        f"Please find attached the invoice for {p.get('brand_name','')} {p.get('project_name','')} Project.\n\n"
                                        f"Amount: {p.get('currency','USD')} {p.get('amount',0):,.2f}\n"
                                        f"Invoice No: {p.get('project_code','')}\n\n"
                                        f"Please review at your convenience and let us know if you have any questions.\n"
                                        f"Thank you for your kind attention.")
                                st.session_state['just_approved'] = {
                                    'name': f"{p.get('brand_name','')}-{m}-invoice.pdf",
                                    'path': tempfile.mktemp(suffix='.pdf'),
                                    'brand': p.get('brand_name',''),
                                    'code': p.get('project_code',''),
                                    'email_subj': subj, 'email_body': body,
                                }
                                _gen_stamped_only(p, st.session_state['just_approved']['path'])
                                st.rerun()
                            except Exception as e: st.error(f"失败: {e}")
                    reject_reason = st.text_input("驳回原因", key=f"rej_reason_{p['id']}",
                                                  placeholder="请填写驳回原因")
                    if st.button("❌ 驳回", key=f"no_{p['id']}", use_container_width=True):
                        if not reject_reason.strip():
                            st.error("请填写驳回原因")
                        else:
                            reject_project(p['id'], user['id'], reject_reason.strip())
                            st.warning("已驳回"); st.rerun()
    else:
        st.success("✅ 没有需要审核的项目")

    # Show just-approved banner with download + email
    if 'just_approved' in st.session_state and st.session_state.get('just_approved'):
        ja = st.session_state['just_approved']
        st.divider()
        st.success(f"✅ 审核通过！{ja['brand']} ({ja['code']})")
        col_dl, col_email = st.columns([1, 2])
        with col_dl:
            if os.path.exists(ja['path']):
                with open(ja['path'], "rb") as f:
                    st.download_button("📥 下载盖章发票PDF", f, file_name=ja['name'],
                                      key="dl_ja", use_container_width=True)
        with col_email:
            with st.expander("📧 邮件文案（发送给客户）", expanded=True):
                st.text_input("主题", value=ja.get('email_subj',''), key="ja_subj")
                st.text_area("正文", value=ja.get('email_body',''), height=180, key="ja_body")
        if st.button("✅ 已处理"):
            st.session_state['just_approved'] = None; st.rerun()

    # Rejected projects
    all_p2 = get_projects(limit=200)
    rejected_list = [p for p in all_p2 if p.get('status')=='rejected']
    if rejected_list:
        st.divider()
        st.subheader(f"❌ 已驳回项目（{len(rejected_list)}个）")
        for p in rejected_list:
            with st.container(border=True):
                c1, c2 = st.columns([4, 1])
                with c1:
                    reason = p.get('rejection_reason','') or '（未填写原因）'
                    st.write(f"**{p.get('brand_name','')}** — {p.get('project_code','') or '待分配'}")
                    st.caption(f"驳回原因：{reason}  |  {p.get('client_short','')}  |  {(p.get('created_at','') or '')[:10]}")
                with c2:
                    if st.button("📝 重新编辑", key=f"reedit_{p['id']}", use_container_width=True):
                        st.session_state['edit_project_id'] = p['id']
                        st.session_state.page = "generate"; st.rerun()

    # Approved projects with download
    st.divider()
    all_p = get_projects(limit=100)
    approved_list = [p for p in all_p if p.get('status')=='approved']
    if approved_list:
        st.subheader(f"✅ 已通过项目（{len(approved_list)}个）")
        for p in approved_list[:20]:
            pid2 = p['id']
            name = (p.get('project_name','') or p.get('brand_name',''))[:60]
            code = p.get('project_code','')

            c1,c2,c3 = st.columns([5,1,1])
            with c1:
                st.write(f"**{name}**  |  {code}")
            with c2:
                try:
                    stamped_path = tempfile.mktemp(suffix='.pdf')
                    _gen_stamped_only(p, stamped_path)
                    ms = code[8:10] if len(code)>=15 else ''
                    M = {'01':'Jan','02':'Feb','03':'Mar','04':'Apr','05':'May','06':'Jun',
                         '07':'Jul','08':'Aug','09':'Sep','10':'Oct','11':'Nov','12':'Dec'}
                    with open(stamped_path, 'rb') as f:
                        st.download_button("📥 下载", f,
                                          file_name=f"{p.get('brand_name','')}-{M.get(ms,'')}-invoice.pdf",
                                          key=f"stamped6_{pid2}", use_container_width=True)
                except: pass
            with c3:
                dl_key = f"dl_{pid2}"
                if dl_key not in st.session_state: st.session_state[dl_key] = False
                if st.button("✅已下载" if st.session_state[dl_key] else "☐未下载", key=f"dlb5_{pid2}", use_container_width=True):
                    st.session_state[dl_key] = not st.session_state[dl_key]
                    st.rerun()
    else:
        st.info("暂无已通过的项目")


def _show_invoice_preview(p):
    """Show a preview of invoice content inline."""
    client = get_client_by_id(p.get('client_id')) or {}
    cur = p.get('currency','USD')
    amt = p.get('amount',0)

    st.markdown(f"""
    <div style="border:1px solid #ddd;border-radius:8px;padding:12px;margin:8px 0;background:#fafafa">
    <b>📄 Invoice 预览</b><br>
    <table style="width:100%;font-size:13px;border-collapse:collapse">
    <tr><td style="padding:3px 8px;color:#888">项目</td><td>{p.get('project_name','')}</td></tr>
    <tr><td style="padding:3px 8px;color:#888">编号</td><td>{p.get('project_code','')}</td></tr>
    <tr><td style="padding:3px 8px;color:#888">客户</td><td>{client.get('full_name','')}</td></tr>
    <tr><td style="padding:3px 8px;color:#888">金额</td><td><b>{cur} {amt:,.2f}</b></td></tr>
    <tr><td style="padding:3px 8px;color:#888">到期日</td><td>{str(p.get('due_date',''))[:10]}</td></tr>
    </table>
    </div>
    """, unsafe_allow_html=True)


def _gen_invoice_dl(p):
    """Generate invoice download button for approval preview."""
    try:
        import openpyxl as xl
        from utils.generate import TEMPLATE_DIR as TD
        from utils.database import generate_project_code, get_connection
        client=get_client_by_id(p.get('client_id')) or {}
        if not client: return

        # Auto-assign code if empty (preview before approval)
        code = p.get('project_code','')
        if not code:
            code = generate_project_code(datetime.now().strftime('%Y-%m-%d'))
            get_connection().table("projects").update({"project_code": code}).eq("id", p['id']).execute()
            p['project_code'] = code

        wb=xl.load_workbook(os.path.join(TD,"Invoice-Template.xlsx")); ws=wb.active
        ws['C3']=f"{p.get('brand_name','')} – {p.get('total_posts','')} CONTENT PACKAGE"
        ws['C7']=client.get('full_name',''); ws['C8']=client.get('address','')
        ws['C9']=client.get('contact',''); ws['C10']=client.get('phone') or ''
        ws['C11']=client.get('email') or ''
        ws['E8']=code; ws['E11']=code
        ws['D15']=p.get('amount',0); ws['E15']=1; ws['G15']=p.get('amount',0)
        ws['E9']=_fmt_date_val(datetime.now())
        ws['E10']=_fmt_date_val(p.get('due_date'))
        _set_c16(ws, p.get("content_type",""))
        _write_c18(ws, p.get('amount',0), p.get('currency','USD'))
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        st.download_button("📥 下载Invoice", buf, file_name=f"{p.get('brand_name','')}-invoice.xlsx",
                          key=f"invdl_{p['id']}", use_container_width=True,
                          mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except: pass


def _set_c16(ws, content_type):
    """Set C16 service label based on invoice type."""
    ct = content_type or ''
    if '前款' in ct: label = "項目【服务】前款"
    elif '中款' in ct: label = "項目【服务】中款"
    elif '后款' in ct or '尾款' in ct: label = "項目【服务】尾款"
    elif '全款' in ct: label = "項目【服务】全款"
    elif '样品费' in ct: label = "項目【样品费】报销"
    elif '差旅费' in ct: label = "項目【差旅费】报销"
    elif '第' in ct and '次' in ct: label = "項目【服务】分期款"
    else: label = "項目【服务】款"
    ws['C16'] = f"{label}\nltem \"Service'"


def _write_c18(ws, amount, currency):
    """Write C18 with Chinese uppercase + English amount."""
    from utils.generate import _amount_chinese
    cl = "RMB" if currency=="RMB" else "USD"
    cn = _amount_chinese(amount, currency)
    ws['C18'] = f"總付款金額為{cn}\nFull payment of {cl} {amount:,.2f}"


def _gen_stamped_only(p, output_path):
    """Generate stamped PDF without approving (for re-download)."""
    import openpyxl as xl
    from utils.pdf_utils import generate_stamped_pdf
    from utils.generate import TEMPLATE_DIR as TD
    from utils.database import generate_project_code, get_connection

    # Safety: auto-assign code if somehow still empty
    code = p.get('project_code','')
    if not code:
        code = generate_project_code(datetime.now().strftime('%Y-%m-%d'))
        get_connection().table("projects").update({"project_code": code}).eq("id", p['id']).execute()

    client=get_client_by_id(p.get('client_id')) or {}
    wb=xl.load_workbook(os.path.join(TD,"Invoice-Template.xlsx")); ws=wb.active
    ws['C3']=f"{p.get('brand_name','')} – {p.get('total_posts','')} CONTENT PACKAGE"
    ws['C7']=client.get('full_name',''); ws['C8']=client.get('address','')
    ws['C9']=client.get('contact',''); ws['C10']=client.get('phone') or ''
    ws['C11']=client.get('email') or ''; ws['E8']=code
    ws['E9']=_fmt_date_val(datetime.now()); ws['E10']=_fmt_date_val(p.get('due_date'))
    ws['E11']=code; ws['D15']=p.get('amount',0); ws['E15']=1; ws['G15']=p.get('amount',0)
    _set_c16(ws, p.get("content_type",""))
    _write_c18(ws, p.get('amount',0), p.get('currency','USD'))
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f: f.write(buf.read()); xlsx_path=f.name
    is_inf = client.get('short_name','') == 'Infinix'
    generate_stamped_pdf(xlsx_path, output_path, add_signature=is_inf)
    try: os.unlink(xlsx_path)
    except: pass


def _regen_and_approve(p, user_id):
    """Regenerate stamped invoice PDF and approve."""
    import openpyxl as xl
    from utils.pdf_utils import generate_stamped_pdf
    from utils.generate import TEMPLATE_DIR as TD
    from utils.database import generate_project_code, get_connection

    # Auto-assign project code based on today if still empty
    code = p.get('project_code','')
    if not code:
        code = generate_project_code(datetime.now().strftime('%Y-%m-%d'))
        get_connection().table("projects").update({"project_code": code}).eq("id", p['id']).execute()
        p['project_code'] = code

    client=get_client_by_id(p.get('client_id')) or {}
    wb=xl.load_workbook(os.path.join(TD,"Invoice-Template.xlsx")); ws=wb.active
    ws['C3']=f"{p.get('brand_name','')} – {p.get('total_posts','')} CONTENT PACKAGE"
    ws['C7']=client.get('full_name',''); ws['C8']=client.get('address','')
    ws['C9']=client.get('contact','')
    ws['C10']=client.get('phone') if client.get('phone') and client['phone']!='（待补充）' else None
    ws['C11']=client.get('email') if client.get('email') and client['email']!='（待补充）' else None
    ws['E8']=code; ws['E9']=_fmt_date_val(datetime.now())
    ws['E10']=_fmt_date_val(p.get('due_date'))
    ws['E11']=code; ws['D15']=p.get('amount',0)
    ws['E15']=1; ws['G15']=p.get('amount',0)
    _set_c16(ws, p.get("content_type",""))
    _write_c18(ws, p.get('amount',0), p.get('currency','USD'))
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    xlsx_path=tempfile.mktemp(suffix='.xlsx')
    with open(xlsx_path,'wb') as f: f.write(buf.read())
    stamped_path=tempfile.mktemp(suffix='.pdf')
    is_inf = client.get('short_name','') == 'Infinix'
    generate_stamped_pdf(xlsx_path, stamped_path, add_signature=is_inf)
    approve_project(p['id'], user_id, stamped_path)
    try: os.unlink(xlsx_path)
    except: pass
