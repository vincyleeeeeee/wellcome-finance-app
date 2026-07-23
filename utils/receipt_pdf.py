"""Generate cash receipt PDF by rendering the Excel as an image, then stamping."""

import os
import io
import random
from datetime import datetime
from PIL import Image, ImageDraw, ImageFont
import openpyxl
from openpyxl.utils import get_column_letter
import pypdf
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.utils import ImageReader
import numpy as np


def generate_receipt_pdf(client: dict, receipt_data: dict, output_path: str = None) -> str:
    """Generate receipt PDF from Excel template rendered as image."""
    import tempfile
    if output_path is None:
        output_path = tempfile.mktemp(suffix=".pdf")

    app_dir = os.path.dirname(os.path.dirname(__file__))
    tmpl = os.path.join(app_dir, "templates", "Cash-Receipt-Template.xlsx")
    if not os.path.exists(tmpl):
        tmpl = "/Users/vincy/Documents/Wellcome/项目/_模板/Cash-Receipt-Template.xlsx"

    wb = openpyxl.load_workbook(tmpl)
    ws = wb.active

    c = client
    rd = receipt_data
    currency = rd.get('currency', 'USD')
    amount = rd.get('payment_amount', rd.get('amount', 0))

    # Fill template
    ws['C3'] = rd.get('project_name', '')
    ws['C4'] = rd.get('project_date', '')
    ws['C5'] = rd.get('venue', '')
    ws['C7'] = c.get('full_name', '')
    ws['C8'] = c.get('address', '')
    ws['C9'] = c.get('contact', '')
    ws['C10'] = c.get('phone') if c.get('phone') and c['phone'] != '（待补充）' else ''
    ws['C11'] = c.get('email') if c.get('email') and c['email'] != '（待补充）' else ''
    ws['E3'] = rd.get('issuer_name', 'Mr. Terry.Su')
    ws['E8'] = rd.get('project_code', '')
    ws['E9'] = _dt(rd.get('payment_date'))
    ws['E10'] = _dt(rd.get('gained_date'))
    ws['E11'] = rd.get('payment_method', 'BANK')
    cl = "RMB" if currency == "RMB" else "USD"
    ws['C13'] = (f"Received From  WELLCOME (INTERNATIONAL) LIMITED    "
                 f"The amount of  {cl}{amount:,.2f}\n"
                 f"For the {rd.get('project_name', '')}  Project")
    d = rd.get('gained_date', datetime.now())
    ds = d.strftime('%Y/%m/%d') if isinstance(d, datetime) else str(d)[:10]
    ws['D16'] = f"Name：\n\nDate：{ds}\n\nSignature：\n"

    # Render xlsx to image
    img = _render_sheet(ws)
    if img is None:
        img = _render_sheet_simple(ws)

    # Stamp overlay
    stamped = _add_stamp(img, at_name_area=True)

    # Save as PDF
    stamped.convert("RGB").save(output_path, "PDF")
    return output_path


def _dt(val):
    if val is None: return ''
    if isinstance(val, datetime): return val.strftime('%Y-%m-%d')
    return str(val)[:10]


def _render_sheet(ws) -> Image.Image:
    """Render worksheet as an image."""
    # Get used range
    min_r, max_r = 1, ws.max_row
    min_c, max_c = 1, ws.max_column

    # Calculate column widths in pixels (1 unit ≈ 7px at 96dpi)
    col_widths_px = {}
    for c in range(min_c, max_c + 1):
        letter = get_column_letter(c)
        w = ws.column_dimensions[letter].width
        col_widths_px[c] = int((w or 10) * 8)

    # Calculate row heights in pixels
    row_heights_px = {}
    for r in range(min_r, max_r + 1):
        h = ws.row_dimensions[r].height
        row_heights_px[r] = int((h or 20) * 1.2)

    total_w = sum(col_widths_px.values()) + 40
    total_h = sum(row_heights_px.values()) + 40

    if total_w <= 0 or total_h <= 0 or total_w > 5000 or total_h > 10000:
        return None

    img = Image.new('RGB', (total_w, total_h), 'white')
    draw = ImageDraw.Draw(img)

    # Try to load fonts
    font_normal = _get_font(11)
    font_small = _get_font(9)
    font_title = _get_font(14)

    # Merge info
    merged = {}
    for mc in ws.merged_cells.ranges:
        for r in range(mc.min_row, mc.max_row + 1):
            for c in range(mc.min_col, mc.max_col + 1):
                merged[(r, c)] = (mc.min_row, mc.min_col)

    y = 20
    for r in range(min_r, max_r + 1):
        x = 20
        rh = row_heights_px.get(r, 24)
        for c in range(min_c, max_c + 1):
            cw = col_widths_px.get(c, 80)

            # Check if this cell is merged into another
            if (r, c) in merged:
                src = merged[(r, c)]
                if src != (r, c):
                    x += cw
                    continue

            cell = ws.cell(r, c)
            val = cell.value
            if val is not None and str(val).strip():
                txt = str(val).replace('\n', ' ')
                font = font_title if r == 1 else (font_small if _is_label(r, c) else font_normal)
                fill = (0, 0, 0) if r > 1 else (0, 0, 0)

                # Check if this cell spans multiple rows/cols
                span_r, span_c = rh, cw
                for mc in ws.merged_cells.ranges:
                    if mc.min_row == r and mc.min_col == c:
                        for sr in range(mc.min_row, mc.max_row + 1):
                            if sr > r:
                                span_r += row_heights_px.get(sr, 24)

                draw.text((x + 2, y + 2), txt[:100], fill=fill, font=font)

            # Cell border
            draw.rectangle([x, y, x + cw, y + rh], outline=(200, 200, 200), width=1)
            x += cw
        y += rh

    return img


def _render_sheet_simple(ws) -> Image.Image:
    """Simple fallback renderer."""
    lines = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        row_texts = []
        for cell in row:
            if cell.value and str(cell.value).strip() and str(cell.value).strip() != '.':
                txt = str(cell.value).replace('\n', ' | ')[:150]
                row_texts.append(txt)
        if row_texts:
            lines.append('  '.join(row_texts))

    if not lines:
        return Image.new('RGB', (800, 600), 'white')

    img = Image.new('RGB', (1200, 40 + len(lines) * 28), 'white')
    draw = ImageDraw.Draw(img)
    font = _get_font(12)

    y = 20
    for line in lines:
        draw.text((20, y), line, fill=(0, 0, 0), font=font)
        y += 28

    return img


def _is_label(r, c):
    """Check if cell is a gray label (column A or D)."""
    return get_column_letter(c) in ('A', 'D')


def _get_font(size):
    """Try to load a font, fall back to default."""
    font_paths = [
        '/System/Library/Fonts/PingFang.ttc',
        '/System/Library/Fonts/STHeiti Light.ttc',
        '/System/Library/Fonts/Helvetica.ttc',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
    ]
    for fp in font_paths:
        try:
            return ImageFont.truetype(fp, size)
        except (IOError, OSError):
            continue
    return ImageFont.load_default()


def _add_stamp(img: Image.Image, at_name_area: bool = False) -> Image.Image:
    """Overlay stamp image. If at_name_area, place near 'Name' signature (left side)."""
    stamp_paths = [
        os.path.join(os.path.dirname(os.path.dirname(__file__)), "stamp", "stamp_500.png"),
        os.path.join(os.path.dirname(os.path.dirname(__file__)), "stamp", "stamp_final.png"),
        "/Users/vincy/Documents/Wellcome/invoice-app/stamp/stamp_500.png",
    ]
    stamp_file = None
    for p in stamp_paths:
        if os.path.exists(p):
            stamp_file = p
            break
    if not stamp_file:
        return img

    stamp = Image.open(stamp_file).convert("RGBA")
    pw, ph = img.size
    stamp_w = int(pw * 0.20)
    ratio = stamp_w / stamp.width
    stamp_h = int(stamp.height * ratio)

    if at_name_area:
        # Place stamp near the Name area (left side, ~80% down)
        x = int(pw * 0.15) + random.randint(-20, 20)
        y = int(ph * 0.65) + random.randint(-20, 20)
    else:
        # Bottom right
        mx = int(pw * 0.04) + random.randint(-15, 15)
        my = int(ph * 0.06) + random.randint(-10, 10)
        x = pw - stamp_w - mx
        y = ph - stamp_h - my

    stamp_r = stamp.resize((stamp_w, stamp_h), Image.LANCZOS)
    rgba = img.convert("RGBA")
    rgba.paste(stamp_r, (x, y), stamp_r)
    return rgba
