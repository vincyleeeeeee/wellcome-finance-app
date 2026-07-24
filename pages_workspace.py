"""Project Workspace — unified project lifecycle management."""

import streamlit as st
import os, io, json, base64, tempfile
from datetime import datetime
from utils.database import (
    get_projects, get_clients, get_client_by_id, get_client_by_short_name,
    get_project_by_id, save_project, submit_for_approval, get_connection
)
from utils.generate import (
    generate_confirmation_letter, generate_invoice,
    generate_email_confirmation, generate_email_invoice,
    TEMPLATE_DIR
)
from utils.pdf_utils import generate_stamped_pdf
import openpyxl

STAGES = {
    'draft': ('📝 基本信息', '可生成确认函'),
    'confirmation_sent': ('📨 确认函已发', '等待客户盖章'),
    'pending': ('⏳ 待审核', '财务审核中'),
    'approved': ('✅ 已开发票', '等待收款'),
}

ACTIONS = {
    'draft': ['generate_confirmation', 'edit'],
    'confirmation_sent': ['upload_stamped', 'edit'],
    'pending': [],
    'approved': ['mark_paid', 'generate_receipt'],
    'rejected': ['edit', 'resubmit'],
}


def page_workspace():
    st.title("📝 项目工作台")
    user = st.session_state.user

    tab1, tab2 = st.tabs(["📋 我的项目", "➕ 新建项目"])

    with tab2:
        _new_project_form(user)

    with tab1:
        _project_list(user)


def _new_project_form(user):
    st.subheader("新建项目（填写一次，后续自动生成各文件）")
    clients = get_clients()
    client_names = [c['short_name'] for c in clients]
    cmap = {c['short_name']: c for c in clients}

    col1, col2 = st.columns(2)
    with col1:
        sel = st.selectbox("客户简称 *", client_names, key="ws_client")
        c = cmap.get(sel, {})
        if c:
            st.caption(f"{c.get('full_name','')} | {c.get('contact','')}")

        code_month = st.selectbox("编号月份", list(range(1,13)),
                                  index=datetime.now().month-1,
                                  format_func=lambda m: f"{m}月")
        code_year = datetime.now().year
        from utils.database import get_next_code_for_month
        try:
            code = get_next_code_for_month(code_year, code_month)
            project_code = st.text_input("项目编号 *", value=code)
        except:
            project_code = st.text_input("项目编号 *", placeholder="WELL260801001")

        project_name = st.text_input("项目名称 *", placeholder="品牌名 月份 UGC 篇数")
        brand_name = st.text_input("客户品牌名 *")

    with col2:
        currency = st.selectbox("币种", ["USD", "RMB"])
        amount = st.number_input("项目金额 *", min_value=0.0, step=100.0, value=0.0)
        venue = st.text_input("执行地点", value="Bangkok")
        execution_period = st.text_input("执行周期", placeholder="2026年8月")
        shooting_date = st.text_input("预计拍摄时间")
        total_posts = st.text_input("总发布篇数", placeholder="150 POSTS")
        due_date = st.date_input("到期日", value=datetime.now())

    # Cost breakdown
    st.divider()
    st.caption("成本构成")
    RATES = {"USD":7.2, "RMB":1.0, "THB":0.2, "MYR":1.55}
    cost_items = []; total_rmb = 0.0
    ccols = st.columns(4)
    for i, cat in enumerate(["拍摄", "餐饮交通", "发布", "补发"]):
        with ccols[i]:
            if st.checkbox(cat, key=f"ws_{cat}"):
                amt = st.number_input(f"金额", key=f"ws_amt_{cat}", value=0.0, step=100.0)
                cur = st.selectbox("币种", ["RMB","USD","THB","MYR"], key=f"ws_cur_{cat}")
                if amt > 0:
                    total_rmb += amt * RATES.get(cur,1)
                    cost_items.append({"name":cat,"amount":amt,"currency":cur})
    custom = st.text_input("其他成本项", placeholder="如KOL费用")
    if custom:
        c1,c2 = st.columns([2,1])
        with c1:
            camt = st.number_input("其他金额", key="ws_camt", value=0.0, step=100.0)
        with c2:
            ccur = st.selectbox("币种", ["RMB","USD","THB","MYR"], key="ws_ccur")
        if camt > 0:
            total_rmb += camt * RATES.get(ccur,1)
            cost_items.append({"name":custom,"amount":camt,"currency":ccur})
    if total_rmb > 0:
        st.info(f"总成本(RMB): ¥{total_rmb:,.0f}")

    content_type = st.text_input("合作内容", value="UGC铺量")
    platform = st.text_input("发布平台", value="小红书")
    expected_payment_date = st.date_input("预计客户到账时间", value=None)

    if st.button("💾 保存项目信息", type="primary", use_container_width=True):
        if not all([sel, project_code, project_name, brand_name, amount > 0]):
            st.error("请填写所有带 * 的必填项")
        else:
            proj = {
                'client_short': sel, 'project_code': project_code,
                'project_name': project_name, 'brand_name': brand_name,
                'amount': amount, 'currency': currency, 'venue': venue,
                'execution_period': execution_period, 'shooting_date': shooting_date,
                'total_posts': total_posts, 'invoice_date': datetime.now().date(),
                'due_date': due_date, 'content_type': content_type,
                'platform': platform, 'status': 'draft',
                'estimated_cost': total_rmb, 'cost_currency': 'RMB',
                'cost_breakdown': json.dumps(cost_items, ensure_ascii=False) if cost_items else '',
                'expected_payment_date': expected_payment_date.strftime('%Y-%m-%d') if expected_payment_date else None,
                'created_by': user['id'], 'client_id': c.get('id'),
                'invoice_project_name': f"{brand_name} – {total_posts} CONTENT PACKAGE",
            }
            pid = save_project(proj)
            st.success(f"✅ 项目已保存！编号: {project_code}")
            st.rerun()


def _project_list(user):
    projects = get_projects(limit=200)
    show_all = st.checkbox("显示所有项目", value=(user['role'] in ('admin','finance')))
    if not show_all:
        projects = [p for p in projects if p.get('created_by') == user['id']]

    if not projects:
        st.info("暂无项目，点击「➕ 新建项目」创建")
        return

    for p in projects:
        stage_label, stage_hint = STAGES.get(p.get('status',''), ('❓',''))
        status = p.get('status','draft')

        with st.container(border=True):
            # Header row
            c1,c2,c3 = st.columns([4,1,1])
            with c1:
                paid = '💰已到账' if p.get('payment_received') else ''
                feishu = '📋飞书' if p.get('feishu_approved') else ''
                st.markdown(f"**{stage_label}** {paid} {feishu} — **{p.get('brand_name','')}** | {p.get('project_code','')}")
                st.caption(f"{p.get('currency','USD')} {p.get('amount',0):,.2f} | {p.get('client_short','')} | {(p.get('created_at','') or '')[:10]} | {stage_hint}")

            with c2:
                # Action buttons based on stage
                if status == 'draft':
                    if st.button("📄 生成确认函", key=f"gencf_{p['id']}", use_container_width=True):
                        _gen_confirmation(p)
                        st.rerun()
                elif status == 'confirmation_sent':
                    _show_upload_stamped(p)
                elif status == 'pending':
                    st.caption("⏳ 等待财务审核")
                elif status == 'approved':
                    if st.button("📥 下载盖章发票", key=f"dlinv_{p['id']}", use_container_width=True):
                        _dl_stamped_invoice(p)
                    if st.button("🧾 开收据", key=f"dorec_{p['id']}", use_container_width=True):
                        st.session_state['receipt_project_id'] = p['id']
                        st.session_state.page = "receipt"
                        st.rerun()
                elif status == 'rejected':
                    if st.button("✏️ 修改重提", key=f"fix_{p['id']}", use_container_width=True):
                        st.session_state['edit_project_id'] = p['id']
                        st.session_state.page = "generate"
                        st.rerun()

            with c3:
                if p.get('status') == 'approved' and not p.get('payment_received'):
                    if st.button("💰 标记到账", key=f"paidw_{p['id']}", use_container_width=True):
                        get_connection().table("projects").update({
                            "payment_received": True,
                            "received_date": datetime.now().strftime('%Y-%m-%d'),
                        }).eq("id", p['id']).execute()
                        st.rerun()
                elif p.get('status') == 'draft':
                    if st.button("✏️ 编辑", key=f"edw_{p['id']}", use_container_width=True):
                        st.session_state['edit_project_id'] = p['id']
                        st.session_state.page = "generate"
                        st.rerun()

            # Show attachments
            if p.get('stamped_confirmation') and status in ('pending','approved'):
                with st.expander("📎 查看盖章确认函"):
                    try:
                        st.image(base64.b64decode(p['stamped_confirmation']))
                    except: pass


def _gen_confirmation(p):
    """Generate confirmation letter and update status."""
    client = get_client_by_id(p.get('client_id')) or {}
    proj = {
        'project_code': p.get('project_code',''), 'project_name': p.get('project_name',''),
        'brand_name': p.get('brand_name',''), 'venue': p.get('venue',''),
        'execution_period': p.get('execution_period',''),
        'shooting_date': p.get('shooting_date',''), 'total_posts': p.get('total_posts',''),
        'amount': p.get('amount',0), 'application_date': datetime.now().strftime('%b %d, %Y'),
    }
    path = generate_confirmation_letter({'full_name': client.get('full_name',''), 'contact': client.get('contact','')}, proj)
    with open(path, 'rb') as f:
        st.download_button("📥 下载确认函", f, file_name=f"{p.get('brand_name','')}-confirmation-letter.docx",
                          key=f"dl_cf_{p['id']}")
    get_connection().table("projects").update({"status": "confirmation_sent"}).eq("id", p['id']).execute()


def _show_upload_stamped(p):
    """Show upload for stamped confirmation."""
    uploaded = st.file_uploader("📎 上传客户盖章确认函", type=["png","jpg","jpeg","pdf"],
                               key=f"up_{p['id']}")
    if uploaded:
        b64 = base64.b64encode(uploaded.read()).decode()
        get_connection().table("projects").update({
            "stamped_confirmation": b64,
            "status": "draft",  # ready for invoice submission
        }).eq("id", p['id']).execute()
        st.success("已上传！现在可以提交生成发票")
        st.rerun()


def _dl_stamped_invoice(p):
    """Download stamped invoice."""
    import tempfile
    client = get_client_by_id(p.get('client_id')) or {}
    wb = openpyxl.load_workbook(os.path.join(TEMPLATE_DIR, "Invoice-Template.xlsx"))
    ws = wb.active
    ws['C3'] = f"{p.get('brand_name','')} – {p.get('total_posts','')} CONTENT PACKAGE"
    ws['C7'] = client.get('full_name','')
    ws['E8'] = p.get('project_code','')
    ws['D15'] = p.get('amount',0); ws['E15'] = 1; ws['G15'] = p.get('amount',0)
    ws['E10'] = str(p.get('due_date',''))[:10]; ws['E11'] = p.get('project_code','')
    from pages_finance import _write_c18
    _write_c18(ws, p.get('amount',0), p.get('currency','USD'))
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        f.write(buf.read()); xlsx_path = f.name
    stamped_path = tempfile.mktemp(suffix='.pdf')
    generate_stamped_pdf(xlsx_path, stamped_path)
    code = p.get('project_code','')
    month_str = code[6:8] if len(code)>=8 else ''
    MONTHS = {'01':'Jan','02':'Feb','03':'Mar','04':'Apr','05':'May','06':'Jun',
              '07':'Jul','08':'Aug','09':'Sep','10':'Oct','11':'Nov','12':'Dec'}
    fname = f"{p.get('brand_name','')}-{MONTHS.get(month_str,'')}-invoice.pdf"
    with open(stamped_path, 'rb') as ff:
        st.download_button("📥 下载盖章发票", ff, file_name=fname, key=f"ws_dl_{p['id']}")
