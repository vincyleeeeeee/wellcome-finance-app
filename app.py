"""Wellcome Invoice & Confirmation Letter Generator — Streamlit App."""

import streamlit as st
import os
import sys
import base64
from datetime import datetime

# Ensure utils is importable
sys.path.insert(0, os.path.dirname(__file__))
from utils.database import (
    init_db, authenticate, create_user, is_approved, get_connection,
    get_pending_users, approve_user, reject_user, get_all_users,
    get_clients, get_client_by_short_name, get_client_by_id,
    upsert_client, delete_client,
    save_project, get_projects, get_project_by_id,
    submit_for_approval, approve_project, reject_project, get_pending_approvals,
    set_user_role, generate_project_code, get_next_code_for_month
)
from utils.generate import (
    generate_confirmation_letter, generate_invoice,
    generate_email_confirmation, generate_email_invoice,
    generate_cash_receipt
)
from utils.pdf_utils import generate_stamped_pdf
from utils.receipt_pdf import generate_receipt_pdf
from pages_finance import page_overview, page_approval
from pages_workspace import page_workspace
from page_generate_new import page_generate


def _fmt_cost(cost_json: str) -> str:
    """Format cost breakdown JSON for display."""
    if not cost_json:
        return ""
    try:
        import json
        items = json.loads(cost_json)
        return "、".join(f"{i['name']}({i['currency']}{i['amount']:,.0f})" for i in items)
    except Exception:
        return cost_json

# ============================================================
# Page config
# ============================================================
st.set_page_config(
    page_title="Wellcome财务自动化平台",
    page_icon="💰",
    layout="wide"
)

# ============================================================
# Custom CSS for sidebar highlighting
# ============================================================
st.markdown("""
<style>
/* === 字体放大，适配年长用户 === */
html, body, [class*="css"] {
    font-size: 16px !important;
}
div[data-testid="stSidebar"] button {
    font-size: 15px !important;
    padding: 12px 8px !important;
    margin: 3px 0 !important;
}
div[data-testid="stMetricValue"] {
    font-size: 28px !important;
}
div[data-testid="stMetricLabel"] {
    font-size: 14px !important;
}
div[data-testid="stSelectbox"] label, div[data-testid="stTextInput"] label,
div[data-testid="stNumberInput"] label, div[data-testid="stDateInput"] label,
div[data-testid="stCheckbox"] label {
    font-size: 15px !important;
}
input, select, textarea, .stTextInput input, .stNumberInput input {
    font-size: 16px !important;
}
button[kind="primary"] {
    font-size: 16px !important;
    padding: 10px 16px !important;
}
h1, .st-emotion-cache-10trblm {
    font-size: 28px !important;
}
h2, .st-emotion-cache-1v0mbdj {
    font-size: 22px !important;
}
h3 {
    font-size: 18px !important;
}
/* Active sidebar button highlight */
div[data-testid="stSidebar"] button[kind="primary"] {
    background-color: #1a73e8 !important;
    color: white !important;
    border: 2px solid #1557b0 !important;
    font-weight: bold !important;
    font-size: 16px !important;
}
div[data-testid="stSidebar"] button[kind="secondary"] {
    background-color: transparent !important;
    color: #555 !important;
    font-size: 15px !important;
}
</style>
""", unsafe_allow_html=True)

# ============================================================
# Init DB
# ============================================================
init_db()

# ============================================================
# Session state
# ============================================================
if "user" not in st.session_state:
    st.session_state.user = None
if "page" not in st.session_state:
    st.session_state.page = "login"

# Auto-restore session using st.session_state (survives reruns)
# For full page refresh, use Streamlit's cookie via extra_streamlit_components
# Simplified: store user_id in session_state, which persists within a browser session

def _save_session(user_id):
    """Store login in session state and query params for persistence."""
    st.session_state['_logged_in_user_id'] = user_id
    st.session_state['_last_active'] = datetime.now().isoformat()

def _restore_session():
    """Try to restore user from session state."""
    uid = st.session_state.get('_logged_in_user_id')
    last = st.session_state.get('_last_active')
    if uid and last:
        # Session valid for 7 days of inactivity
        try:
            last_dt = datetime.fromisoformat(last)
            if (datetime.now() - last_dt).days < 7:
                result = get_connection().table("users").select("*").eq("id", uid).execute()
                if result.data:
                    return result.data[0]
        except:
            pass
    return None

def _heartbeat():
    """Keep session alive on every interaction."""
    if st.session_state.get('_logged_in_user_id'):
        st.session_state['_last_active'] = datetime.now().isoformat()

def _clear_session():
    for k in ['_logged_in_user_id', '_last_active']:
        if k in st.session_state:
            del st.session_state[k]


def logout():
    _clear_session()
    st.session_state.user = None
    st.session_state.page = "login"


# ============================================================
# Sidebar navigation (when logged in)
# ============================================================
def render_sidebar():
    _heartbeat()
    user = st.session_state.user
    role_labels = {'admin': '🔑 管理员', 'finance': '💰 财务', 'user': '👤 用户'}
    with st.sidebar:
        st.markdown(f"### 👤 {user['username']}")
        st.caption(f"{user['email']}  |  {role_labels.get(user['role'], user['role'])}")

        st.divider()

        if st.button("📝 项目工作台", use_container_width=True,
                     type="primary" if st.session_state.page == "workspace" else "secondary"):
            st.session_state.page = "workspace"

        if st.button("📄 生成文档", use_container_width=True,
                     type="primary" if st.session_state.page == "generate" else "secondary"):
            st.session_state.page = "generate"

        if st.button("👥 客户管理", use_container_width=True,
                     type="primary" if st.session_state.page == "clients" else "secondary"):
            st.session_state.page = "clients"


        # Finance users: simple, clear pages
        if user['role'] in ('finance', 'admin'):
            st.divider()
            st.markdown("**💰 财务专区**")
            if st.button("📊 项目总览", use_container_width=True,
                         type="primary" if st.session_state.page == "overview" else "secondary"):
                st.session_state.page = "overview"
            p = len(get_pending_approvals())
            label = f"⏳ 待审核" + (f" ({p})" if p else "")
            if st.button(label, use_container_width=True,
                         type="primary" if st.session_state.page == "approval" else "secondary"):
                st.session_state.page = "approval"
            if st.button("🧾 开收据", use_container_width=True,
                         type="primary" if st.session_state.page == "receipt" else "secondary"):
                st.session_state.page = "receipt"

        if user['role'] == 'admin':
            pending = len(get_pending_users())
            admin_label = f"🔒 用户审核" + (f" ({pending})" if pending else "")
            if st.button(admin_label, use_container_width=True,
                         type="primary" if st.session_state.page == "admin" else "secondary"):
                st.session_state.page = "admin"

        st.divider()
        st.button("🚪 退出登录", on_click=logout, use_container_width=True)


# ============================================================
# Login / Register page
# ============================================================
def _sync_login_input(field):
    """Sync autofill values into session state."""
    if field == 'email':
        st.session_state['login_email_val'] = st.session_state.get('login_email_field', '')
    else:
        st.session_state['login_password_val'] = st.session_state.get('login_password_field', '')


def page_login():
    st.title("💰 Wellcome财务自动化平台")
    st.caption("确认函 · Invoice · 收据 — 一键生成")

    tab_login, tab_register = st.tabs(["登录", "注册"])

    with tab_login:
        st.subheader("登录")
        with st.form("login_form"):
            email = st.text_input("邮箱", placeholder="your@email.com")
            password = st.text_input("密码", type="password")
            submitted = st.form_submit_button("🔐 登录", type="primary", use_container_width=True)
        if submitted:
            user = authenticate(email.strip(), password.strip())
            if user is None:
                st.error("邮箱或密码错误")
            elif isinstance(user, dict) and '_error' in user:
                st.error(f"服务器连接失败: {user['_error']}")
            elif user['approved'] == 0:
                st.warning("你的账号尚未通过审核，请等待管理员审批")
            else:
                _save_session(user['id'])
                st.session_state.user = user
                st.session_state.page = "workspace"
                st.rerun()

    with tab_register:
        st.subheader("注册")
        st.info("注册后需等待管理员审核通过才能使用")
        new_email = st.text_input("邮箱", key="reg_email", placeholder="your@email.com")
        new_username = st.text_input("用户名", key="reg_username", placeholder="你的名字")
        new_password = st.text_input("密码", type="password", key="reg_password", placeholder="至少6位")
        new_password2 = st.text_input("确认密码", type="password", key="reg_password2", placeholder="再次输入密码")

        if st.button("提交注册", use_container_width=True):
            if not new_email or not new_username or not new_password:
                st.error("请填写所有字段")
            elif new_password != new_password2:
                st.error("两次密码不一致")
            elif len(new_password) < 6:
                st.error("密码至少6位")
            else:
                success, msg = create_user(new_email.strip(), new_username.strip(), new_password.strip())
                if success:
                    st.success(msg)
                else:
                    st.error(msg)


# ============================================================
# Admin: User Approval
# ============================================================
def page_admin():
    st.title("🔒 用户审核")
    user = st.session_state.user
    if user['role'] != 'admin':
        st.error("无权访问")
        return

    st.subheader("待审核用户")
    pending = get_pending_users()
    if not pending:
        st.success("暂无待审核用户 ✅")
    else:
        for u in pending:
            col1, col2, col3 = st.columns([3, 1, 1])
            with col1:
                st.write(f"**{u['username']}**")
                st.caption(f"{u['email']}  |  注册于 {u['created_at']}")
            with col2:
                if st.button("✅ 通过", key=f"approve_{u['id']}"):
                    approve_user(u['id'])
                    st.rerun()
            with col3:
                if st.button("❌ 拒绝", key=f"reject_{u['id']}"):
                    reject_user(u['id'])
                    st.rerun()
            st.divider()

    st.subheader("所有用户")
    users = get_all_users()
    role_map = {'admin': '🔑 管理员', 'finance': '💰 财务', 'user': '👤 用户'}
    for u in users:
        status = "✅" if u['approved'] else "⏳"
        role = role_map.get(u['role'], u['role'])
        col1, col2, col3 = st.columns([3, 1, 1])
        with col1:
            st.write(f"{status} {u['username']} — {u['email']} — {role} — {u['created_at'][:10]}")
        with col2:
            new_role = st.selectbox(
                "角色", ['user', 'finance', 'admin'],
                index=['user', 'finance', 'admin'].index(u['role']) if u['role'] in ['user', 'finance', 'admin'] else 0,
                key=f"role_{u['id']}",
                label_visibility="collapsed"
            )
        with col3:
            if new_role != u['role']:
                if st.button("💾 保存", key=f"save_role_{u['id']}"):
                    set_user_role(u['id'], new_role)
                    st.rerun()


# ============================================================
# Client Management
# ============================================================
def page_clients():
    st.title("👥 客户管理")

    clients = get_clients()
    client_names = [c['short_name'] for c in clients]
    client_map = {c['short_name']: c for c in clients}

    mode = st.radio("操作", ["查看 / 编辑", "新增客户", "删除客户"], horizontal=True)

    if mode == "新增客户":
        _client_form(None)
    elif mode == "删除客户":
        if not client_names:
            st.info("暂无客户")
        else:
            target = st.selectbox("选择要删除的客户", client_names)
            if st.button("🗑️ 确认删除", type="primary"):
                c = client_map[target]
                st.warning(f"确认删除客户「{target}」({c['full_name']})？")
                if st.button("⚠️ 再次确认删除", type="primary"):
                    success, msg = delete_client(target)
                    if success:
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)
    else:  # 查看/编辑
        if not client_names:
            st.info("暂无客户，请先新增")
        else:
            selected = st.selectbox("选择客户", client_names)
            if selected:
                _client_form(client_map[selected])


def _client_form(client: "dict | None"):
    """Render client add/edit form."""
    is_edit = client is not None
    st.subheader("编辑客户" if is_edit else "新增客户")

    short_name = st.text_input("简称 *", value=client['short_name'] if is_edit else "",
                               disabled=is_edit, placeholder="如 AKRA, POP")
    full_name = st.text_input("公司全称 *", value=client['full_name'] if is_edit else "",
                              placeholder="公司完整名称")
    address = st.text_area("公司地址", value=client['address'] if is_edit else "",
                           placeholder="详细地址")
    contact = st.text_input("联系人", value=client['contact'] if is_edit else "",
                            placeholder="联系人姓名")
    phone = st.text_input("电话", value=client['phone'] if is_edit else "",
                          placeholder="(+66) ...")
    email = st.text_input("邮箱", value=client['email'] if is_edit else "",
                          placeholder="contact@company.com")

    btn_label = "💾 保存修改" if is_edit else "➕ 创建客户"
    if st.button(btn_label, type="primary", use_container_width=True):
        if not short_name or not full_name:
            st.error("简称和公司全称为必填项")
        else:
            success, msg = upsert_client(
                short_name, full_name, address, contact, phone, email,
                st.session_state.user['id']
            )
            if success:
                st.success(msg)
                st.rerun()
            else:
                st.error(msg)


# page_generate imported from page_generate_new.py


def _page_generate_placeholder():
    st.title("📄 生成文档")
    user = st.session_state.user

    # === Progress stages ===
    STAGES_LIST = [
        ('info', '📝 填写信息', '填写项目基本信息'),
        ('confirmation', '📄 确认函', '生成并发送确认函给客户'),
        ('stamped', '📎 盖章回传', '客户盖章确认函上传'),
        ('invoice', '🧾 开发票', '提交财务审核，生成盖章发票'),
        ('receipt', '💰 开收据', '客户付款后开收据'),
    ]

    # Check if editing/selecting from workspace
    edit_id = st.session_state.get('edit_project_id')
    edit_data = None
    if edit_id:
        edit_data = get_project_by_id(edit_id)
        if edit_data:
            st.info(f"🔧 正在编辑项目：**{edit_data.get('project_code','')}** {edit_data.get('brand_name','')}")
            if st.button("❌ 取消编辑"):
                del st.session_state['edit_project_id']
                st.rerun()
        else:
            del st.session_state['edit_project_id']

    clients = get_clients()
    if not clients:
        st.warning("暂无客户，请先在「客户管理」中添加")
        return

    client_names = [c['short_name'] for c in clients]
    client_map = {c['short_name']: c for c in clients}

    col_left, col_right = st.columns(2)

    with col_left:
        st.subheader("项目信息")
        # Pre-fill from edit data
        default_client_idx = 0
        if edit_data:
            edit_client = get_client_by_id(edit_data.get('client_id'))
            if edit_client and edit_client.get('short_name') in client_names:
                default_client_idx = client_names.index(edit_client['short_name'])

        selected_client = st.selectbox("客户简称 *", client_names, index=default_client_idx)
        client_info = client_map[selected_client]

        phone_email = client_info['phone'] or client_info['email'] or '—'
        st.caption(f"📋 {client_info['full_name']}  |  联系人: {client_info['contact']}  |  {phone_email}")

        application_date = st.date_input("申请日期 *", value=datetime.now())
        # Month selector for project code
        col_month, col_show = st.columns([1, 2])
        with col_month:
            code_month = st.selectbox("编号月份", list(range(1, 13)),
                                      index=datetime.now().month - 1,
                                      format_func=lambda m: f"{m}月",
                                      help="选月份查看该月最新编号")
            code_year = datetime.now().year
            if code_month < datetime.now().month:
                code_year += 1  # if user selects earlier month, assume next year
        with col_show:
            if edit_data and edit_data.get('project_code'):
                project_code = st.text_input("项目编号 *", value=edit_data['project_code'])
            else:
                try:
                    latest = get_next_code_for_month(code_year, code_month)
                    st.success(f"📝 **{code_month}月** 下一个可用编号：**{latest}**（实时，不会重复）")
                except Exception:
                    latest = f"WELL{code_year % 100:02d}{code_month:02d}01XX"
                project_code = st.text_input("项目编号 *", value=latest, help="自动生成，可直接修改")
        project_name = st.text_input("项目名称 *", value=edit_data.get('project_name','') if edit_data else '',
                                     placeholder="品牌名 – 月份UGC 篇数")
        brand_name = st.text_input("客户品牌名 *", value=edit_data.get('brand_name','') if edit_data else '',
                                   placeholder="品牌的社交媒体名")
        default_cur_idx = 0 if (edit_data.get('currency','USD') == 'USD' if edit_data else True) else 1
        currency = st.selectbox("币种", ["USD", "RMB"], index=default_cur_idx)
        amount = st.number_input("项目金额 *", min_value=0.0, step=100.0,
                                 value=float(edit_data.get('amount',0)) if edit_data else None)

        venue = st.text_input("执行地点", value=edit_data.get('venue','Bangkok') if edit_data else "Bangkok")
        execution_period = st.text_input("执行周期", value=edit_data.get('execution_period','') if edit_data else '',
                                         placeholder="如 July – September 2026")
        shooting_date = st.text_input("预计拍摄时间", value=edit_data.get('shooting_date','') if edit_data else '',
                                      placeholder="如 July 2026")
        total_posts = st.text_input("总发布篇数", value=edit_data.get('total_posts','') if edit_data else '',
                                    placeholder="如 150 PHOTO POSTS")
        due_date = st.date_input("到期日 *", value=datetime.now())

    with col_right:
        st.subheader("额外信息")
        content_type = st.text_input("合作内容", value="UGC铺量",
                                     help="邮件中使用，默认 UGC铺量")
        platform = st.text_input("发布平台", value="小红书",
                                 help="邮件中使用，默认小红书")
        only_invoice = st.checkbox("只要 Invoice（不生成确认函）", value=False)
        only_confirmation = st.checkbox("只要确认函（不生成 Invoice）", value=False)
        feishu_approved = st.checkbox("已在飞书立项（财务审核通过前提）", value=False,
                                     help="勾选后需上传飞书审批截图")
        submit_approval = st.checkbox("生成后提交财务审核", value=not feishu_approved)
        feishu_screenshot = None
        if feishu_approved or submit_approval:
            feishu_screenshot = st.file_uploader(
                "📎 上传飞书项目审批截图 *", type=["png","jpg","jpeg","webp","bmp"],
                accept_multiple_files=False,
                help="支持拖拽、点击上传、粘贴截图"
            )
        expected_payment_date = st.date_input("预计客户到账时间", value=None,
                                              help="客户的预计付款日期，用于财务跟踪")
        # === Cost breakdown with individual amounts ===
        st.caption("成本构成（勾选后填入金额）")
        RATES = {"USD": 7.2, "RMB": 1.0, "THB": 0.2, "MYR": 1.55}

        # Pre-fill cost items from edit data
        edit_cost_map = {}
        if edit_data and edit_data.get('cost_breakdown'):
            try:
                import json
                for item in json.loads(edit_data['cost_breakdown']):
                    edit_cost_map[item['name']] = item
            except: pass

        cost_items_data = []
        total_rmb = 0.0

        cost_cats = ["拍摄", "餐饮交通", "发布", "补发"]
        for cat in cost_cats:
            pre = edit_cost_map.get(cat, {})
            use_cat = st.checkbox(cat, value=(cat in edit_cost_map), key=f"cost_{cat}")
            if use_cat:
                cc1, cc2 = st.columns([2, 1])
                with cc1:
                    amt = st.number_input(f"{cat}金额", min_value=0.0, step=100.0,
                                          value=float(pre.get('amount', 0)) if pre.get('amount') else None, key=f"amt_{cat}")
                with cc2:
                    cur_list = ["RMB", "USD", "THB", "MYR"]
                    cur_idx = cur_list.index(pre.get('currency', 'RMB')) if pre.get('currency') in cur_list else 0
                    cur = st.selectbox("币种", cur_list, index=cur_idx, key=f"cur_{cat}")
                if amt and amt > 0:
                    total_rmb += amt * RATES.get(cur, 1)
                    cost_items_data.append({"name": cat, "amount": amt, "currency": cur})

        # Custom item
        custom_pre = {}
        for k, v in edit_cost_map.items():
            if k not in cost_cats:
                custom_pre = v
                break
        custom_name = st.text_input("其他项名称", value=custom_pre.get('name', ''),
                                    placeholder="如：KOL费用", key="cost_custom_name")
        if custom_name:
            cc1, cc2 = st.columns([2, 1])
            with cc1:
                custom_amt = st.number_input(f"{custom_name}金额", min_value=0.0, step=100.0,
                                             value=float(custom_pre.get('amount', 0)) if custom_pre.get('amount') else None, key="amt_custom")
            with cc2:
                cur_list = ["RMB", "USD", "THB", "MYR"]
                cur_idx = cur_list.index(custom_pre.get('currency', 'RMB')) if custom_pre.get('currency') in cur_list else 0
                custom_cur = st.selectbox("币种", cur_list, index=cur_idx, key="cur_custom")
            if custom_amt and custom_amt > 0:
                total_rmb += custom_amt * RATES.get(custom_cur, 1)
                cost_items_data.append({"name": custom_name, "amount": custom_amt, "currency": custom_cur})

        if total_rmb > 0:
            st.info(f"💰 预估总成本（人民币）：**¥{total_rmb:,.0f}**")

        # Store as structured JSON string
        import json
        cost_breakdown = json.dumps(cost_items_data, ensure_ascii=False) if cost_items_data else ""
        estimated_cost = total_rmb
        cost_currency = "RMB"

        st.divider()

        if st.button("🚀 生成文档", type="primary", use_container_width=True):
            if not all([project_code, project_name, brand_name, amount > 0]):
                st.error("请填写所有带 * 的必填项")
            elif submit_approval and not feishu_screenshot:
                st.error("提交财务审核必须上传飞书项目审批截图")
            else:
                # Build project dict for generation
                proj = {
                    'client_short': selected_client,
                    'project_code': project_code,
                    'project_name': project_name,
                    'brand_name': brand_name,
                    'amount': amount,
                    'currency': currency,
                    'application_date': application_date.strftime("%b %dth, %Y") if application_date else "",
                    'venue': venue,
                    'execution_period': execution_period,
                    'shooting_date': shooting_date,
                    'total_posts': total_posts,
                    'invoice_date': application_date,
                    'due_date': due_date,
                    'content_type': content_type,
                    'platform': platform,
                    'status': 'pending' if submit_approval else 'draft',
                    'estimated_cost': estimated_cost,
                    'cost_currency': cost_currency,
                    'cost_breakdown': cost_breakdown,
                    'feishu_approved': feishu_approved,
                    'feishu_screenshot': base64.b64encode(feishu_screenshot.read()).decode() if feishu_screenshot else '',
                    'expected_payment_date': expected_payment_date.strftime('%Y-%m-%d') if expected_payment_date else None,
                    'created_by': st.session_state.user['id'],
                    'client_id': client_info['id'],
                }
                proj['invoice_project_name'] = f"{brand_name} – {total_posts} CONTENT PACKAGE"

                with st.spinner("正在生成文档..."):
                    files = {}
                    if not only_invoice:
                        conf_path = generate_confirmation_letter(
                            {k: client_info[k] for k in ['full_name', 'contact']},
                            proj
                        )
                        files['确认函'] = conf_path

                    if not only_confirmation:
                        inv_path = generate_invoice(client_info, proj)
                        files['Invoice'] = inv_path

                    # Save or update project record
                    if edit_id:
                        # Update existing project
                        from utils.database import get_connection as _gc
                        _gc().table("projects").update({
                            "project_code": project_code,
                            "project_name": project_name,
                            "client_id": client_info['id'],
                            "brand_name": brand_name,
                            "amount": amount, "currency": currency,
                            "venue": venue, "execution_period": execution_period,
                            "shooting_date": shooting_date, "total_posts": total_posts,
                            "content_type": content_type, "platform": platform,
                            "status": "pending" if submit_approval else "draft",
                            "estimated_cost": estimated_cost, "cost_currency": cost_currency,
                            "cost_breakdown": cost_breakdown,
                            "feishu_approved": feishu_approved,
                            "expected_payment_date": expected_payment_date.strftime('%Y-%m-%d') if expected_payment_date else None,
                        }).eq("id", edit_id).execute()
                        project_id = edit_id
                        del st.session_state['edit_project_id']
                    else:
                        project_id = save_project(proj)

                    # Generate emails
                    subj_conf, body_conf = ("", "")
                    if not only_invoice:
                        subj_conf, body_conf = generate_email_confirmation(proj)
                    subj_inv, body_inv = generate_email_invoice(proj)

                st.success("✅ 文档生成完毕！" + (" 已提交财务审核" if submit_approval else ""))
                st.balloons()

                # Store project id for possible later approval
                st.session_state['last_project_id'] = project_id
                st.session_state['last_inv_path'] = inv_path

                # Show downloaded files
                for name, path in files.items():
                    with open(path, "rb") as f:
                        st.download_button(
                            label=f"📥 下载 {name}",
                            data=f,
                            file_name=os.path.basename(path),
                            mime="application/octet-stream",
                            use_container_width=True
                        )

                # Show email copy
                if not only_invoice:
                    with st.expander("📧 邮件一：确认函（发客户确认）", expanded=True):
                        st.text_input("主题", value=subj_conf, key="email_subj_conf")
                        st.text_area("正文", value=body_conf, height=250, key="email_body_conf")

                with st.expander("📧 邮件二：Invoice（发客户查收）", expanded=not only_invoice):
                    st.text_input("主题", value=subj_inv, key="email_subj_inv")
                    st.text_area("正文", value=body_inv, height=200, key="email_body_inv")

                # Allow late submission if not submitted initially
                if not submit_approval:
                    if st.button("📤 补交财务审核", use_container_width=True):
                        submit_for_approval(project_id)
                        st.success("已提交财务审核")
                        st.rerun()


# ============================================================
# Project History
# ============================================================
def page_history():
    st.title("📋 项目历史")
    user = st.session_state.user

    # Show rejected projects that need attention
    rejected_mine = [p for p in get_projects(limit=200) if p.get('status') == 'rejected' and p.get('created_by') == user['id']]
    if rejected_mine:
        for rp in rejected_mine:
            st.warning(f"⚠️ 你的项目 **{rp.get('brand_name','')}** ({rp.get('project_code','')}) 已被驳回，请修改后重新提交")
            if st.button("📤 修改并重新提交", key=f"resubmit_{rp['id']}"):
                st.session_state['edit_project_id'] = rp['id']
                st.session_state.page = "generate"
                st.rerun()

    # Filter: show my projects or all
    show_all = st.checkbox("显示所有项目", value=(user['role'] in ('admin', 'finance')))
    projects = get_projects(limit=200)
    if not show_all:
        projects = [p for p in projects if p.get('created_by') == user['id']]

    if not projects:
        st.info("暂无项目记录")
        return

    status_map = {'draft': '草稿', 'pending': '待审核', 'approved': '✅ 已通过', 'rejected': '❌ 已驳回'}
    closure_map = {'active': '🟢 进行中', 'pending_payment': '🟡 待付款', 'closed': '🔵 已结案'}

    # Track selections
    if '_selected_projects' not in st.session_state:
        st.session_state['_selected_projects'] = set()

    # Batch action buttons
    sel_count = len(st.session_state['_selected_projects'])
    if sel_count > 0:
        st.info(f"已选择 {sel_count} 个项目")
        bc1, bc2, bc3 = st.columns(3)
        with bc1:
            if st.button("📤 批量提交审核", use_container_width=True):
                for pid in list(st.session_state['_selected_projects']):
                    submit_for_approval(pid)
                st.session_state['_selected_projects'] = set()
                st.success(f"已提交 {sel_count} 个项目")
                st.rerun()
        with bc2:
            if st.button("🗑️ 批量删除", use_container_width=True):
                for pid in list(st.session_state['_selected_projects']):
                    from utils.database import get_connection as _gc
                    _gc().table("projects").delete().eq("id", pid).execute()
                st.session_state['_selected_projects'] = set()
                st.success(f"已删除 {sel_count} 个项目")
                st.rerun()
        with bc3:
            if st.button("❌ 清除选择", use_container_width=True):
                st.session_state['_selected_projects'] = set()
                st.rerun()

    for p in projects:
        status_label = status_map.get(p.get('status'), p.get('status', '?'))
        closure = p.get('closure_status', 'active') or 'active'
        closure_label = closure_map.get(closure, closure)

        with st.container(border=True):
            # Checkbox + basic info
            cc0, cc1, cc2 = st.columns([0.5, 3, 1])
            pid = p['id']
            sel_key = f"sel_{pid}"
            with cc0:
                selected = st.checkbox("", key=sel_key, label_visibility="collapsed",
                                       value=pid in st.session_state['_selected_projects'])
                if selected:
                    st.session_state['_selected_projects'].add(pid)
                else:
                    st.session_state['_selected_projects'].discard(pid)
            with cc1:
                st.write(f"{status_label} {closure_label} **{p.get('brand_name','?')}** — {p.get('project_code','?')}")
                info_parts = [f"金额: {p.get('currency','USD')} {p.get('amount',0):,.2f}"]
                if p.get('estimated_cost'):
                    info_parts.append(f"成本: {p.get('cost_currency','USD')} {p.get('estimated_cost',0):,.0f}")
                info_parts.append(f"提交: {p.get('created_at','')[:10]}")
                if p.get('expected_payment_date'):
                    info_parts.append(f"预计付款: {p['expected_payment_date']}")
                st.caption(" | ".join(info_parts))

            with cc2:
                if p.get('status') == 'approved' and p.get('stamped_pdf_path'):
                    stamped = p['stamped_pdf_path']
                    if os.path.exists(stamped):
                        with open(stamped, "rb") as f:
                            code_p = p.get('project_code','')
                            ms = code_p[6:8] if len(code_p)>=8 else ''
                            M = {'01':'Jan','02':'Feb','03':'Mar','04':'Apr','05':'May','06':'Jun','07':'Jul','08':'Aug','09':'Sep','10':'Oct','11':'Nov','12':'Dec'}
                            fname = f"{p.get('brand_name','')}-{M.get(ms,'')}-invoice.pdf"
                            st.download_button("📥 盖章PDF", f, file_name=fname,
                                             key=f"hist_stamped_{p['id']}", use_container_width=True)
                            # Email template
                            with st.expander("📧 邮件文案", expanded=False):
                                code_p2 = p.get('project_code','')
                                ms2 = code_p2[6:8] if len(code_p2)>=8 else ''
                                M2 = {'01':'Jan','02':'Feb','03':'Mar','04':'Apr','05':'May','06':'Jun','07':'Jul','08':'Aug','09':'Sep','10':'Oct','11':'Nov','12':'Dec'}
                                subj = f"Invoice for {p.get('brand_name','')} {M2.get(ms2,'')} Campaign / {code_p2}"
                                body = f"Dear all,\n\nPlease find attached the invoice for {p.get('brand_name','')} {p.get('project_name','')} Project.\n\nAmount: {p.get('currency','USD')} {p.get('amount',0):,.2f}\nInvoice No: {code_p2}\n\nPlease review at your convenience and let us know if you have any questions.\nThank you for your kind attention."
                                st.text_input("主题", value=subj, key=f"histsubj_{p['id']}")
                                st.text_area("正文", value=body, height=150, key=f"histbody_{p['id']}")
                elif p.get('status') in ('draft', 'rejected') and user['id'] == p.get('created_by'):
                    if st.button("📤 提交审核", key=f"submit_hist_{p['id']}", use_container_width=True):
                        submit_for_approval(p['id'])
                        st.success(f"已提交: {p.get('project_code','')}")
                        st.rerun()

            # Row 2: editable fields (visible to project owner or finance/admin)
            if user['id'] == p.get('created_by') or user['role'] in ('finance', 'admin'):
                with st.expander("✏️ 编辑项目状态", expanded=False):
                    ce1, ce2, ce3 = st.columns(3)
                    with ce1:
                        new_closure = st.selectbox(
                            "结案状态", ['active', 'pending_payment', 'closed'],
                            index=['active', 'pending_payment', 'closed'].index(closure),
                            format_func=lambda x: closure_map.get(x, x),
                            key=f"closure_{p['id']}"
                        )
                    with ce2:
                        curr_ed = p.get('expected_payment_date')
                        new_ed = st.date_input("预计客户付款时间",
                                              value=datetime.strptime(curr_ed, '%Y-%m-%d') if curr_ed else None,
                                              key=f"expay_{p['id']}")
                    with ce3:
                        if st.button("💾 保存", key=f"save_hist_{p['id']}", use_container_width=True):
                            sb = get_connection()
                            updates = {"closure_status": new_closure}
                            if new_ed:
                                updates["expected_payment_date"] = new_ed.strftime('%Y-%m-%d')
                            if new_closure == 'closed':
                                updates["actual_payment_date"] = datetime.now().strftime('%Y-%m-%d')
                            sb.table("projects").update(updates).eq("id", p['id']).execute()
                            st.success("已更新！")
                            st.rerun()

                    # Finance: send reminder
                    if user['role'] in ('finance', 'admin') and not p.get('reminder_sent'):
                        with ce3:
                            if st.button("🔔 提醒负责人", key=f"remind_{p['id']}", use_container_width=True):
                                note = f"财务提醒：项目 {p['project_code']} 请更新客户付款时间及结案状态"
                                sb = get_connection()
                                sb.table("projects").update({
                                    "reminder_sent": True,
                                    "reminder_note": note,
                                    "reminder_sent_at": datetime.now().isoformat()
                                }).eq("id", p['id']).execute()
                                st.success("已提醒！")
                                st.rerun()
                    elif p.get('reminder_sent'):
                        st.info(f"🔔 已提醒 | {p.get('reminder_note','')[:80]}")


# ============================================================
# Finance Approval Page
# ============================================================
def _regenerate_invoice_xlsx(client: dict, project: dict) -> bytes:
    """Regenerate invoice xlsx from data, return bytes."""
    import io
    from utils.generate import TEMPLATE_DIR
    import openpyxl

    wb = openpyxl.load_workbook(os.path.join(TEMPLATE_DIR, "Invoice-Template.xlsx"))
    ws = wb.active
    currency = project.get('currency', 'USD')
    amount = project.get('amount', 0)

    ws['C3'] = f"{project.get('brand_name','')} – {project.get('total_posts','')} CONTENT PACKAGE"
    ws['C4'] = project.get('execution_period', '')
    ws['C5'] = project.get('venue', '')
    ws['C7'] = client.get('full_name', '')
    ws['C8'] = client.get('address', '')
    ws['C9'] = client.get('contact', '')
    ws['C10'] = client.get('phone') if client.get('phone') and client['phone'] != '（待补充）' else None
    ws['C11'] = client.get('email') if client.get('email') and client['email'] != '（待补充）' else None
    ws['E8'] = project.get('project_code', '')
    ws['E9'] = project.get('invoice_date', '')
    ws['E10'] = project.get('due_date', '')
    ws['E11'] = project.get('project_code', '')
    ws['D15'] = amount; ws['E15'] = 1; ws['G15'] = amount
    ws['C18'] = f"Payment of {currency} {amount:,.2f}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _regenerate_stamped_pdf_from_data(client: dict, project: dict) -> bytes:
    """Generate stamped invoice PDF from data."""
    import io, tempfile
    xlsx_bytes = _regenerate_invoice_xlsx(client, project)
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        f.write(xlsx_bytes)
        xlsx_path = f.name
    pdf_path = tempfile.mktemp(suffix='.pdf')
    try:
        from utils.pdf_utils import generate_stamped_pdf
        generate_stamped_pdf(xlsx_path, pdf_path)
        with open(pdf_path, 'rb') as f:
            return f.read()
    finally:
        try: os.unlink(xlsx_path)
        except: pass
        try: os.unlink(pdf_path)
        except: pass


def page_finance():
    st.title("💰 财务审核")
    user = st.session_state.user
    if user['role'] not in ('finance', 'admin'):
        st.error("无权访问，仅财务角色可审核")
        return

    pending = get_pending_approvals()
    if not pending:
        st.success("暂无待审核项目 ✅")
        return

    st.subheader(f"待审核 ({len(pending)} 个)")

    for p in pending:
        with st.container(border=True):
            col1, col2 = st.columns([3, 1])
            with col1:
                st.markdown(f"**{p['brand_name']}** — {p['project_name']}")
                cost_info = ""
                if p.get('estimated_cost'):
                    cost_info = f" | 成本: {p.get('cost_currency','USD')} {p['estimated_cost']:,.2f}"
                    cb = _fmt_cost(p.get('cost_breakdown', '') or '')
                    if cb:
                        cost_info += f" ({cb[:80]})"
                feishu_badge = " ✅飞书已立项" if p.get('feishu_approved') else " ⚠️未确认飞书立项"
                st.caption(f"编号: {p['project_code']} | 客户: {p['client_short']} | "
                          f"金额: {p.get('currency','USD')} {p['amount']:,.2f}{cost_info} | "
                          f"提交人: {p.get('created_by_name','?')} | "
                          f"提交时间: {p['created_at']} | "
                          f"{feishu_badge}")

                # Regenerate invoice file for download (works on cloud)
                try:
                    client = get_client_by_id(p['client_id']) or {}
                    if client:
                        inv_bytes = _regenerate_invoice_xlsx(client, p)
                        brand = p.get('brand_name', 'project')
                        fname = f"{brand}-{p.get('project_code','invoice')}.xlsx"
                        st.download_button(
                            f"📥 下载 Invoice (审核用)",
                            inv_bytes, file_name=fname,
                            key=f"dl_{p['id']}",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    else:
                        st.caption("⚠️ 客户信息缺失")
                except Exception as e:
                    st.caption(f"生成失败: {e}")

            with col2:
                st.write("")  # spacer
                st.write("")
                if st.button("✅ 通过", key=f"approve_{p['id']}", use_container_width=True, type="primary"):
                    client = get_client_by_id(p.get('client_id')) or {}
                    if not client:
                        st.error("找不到客户信息")
                    else:
                        with st.spinner("正在生成盖章 PDF..."):
                            try:
                                pdf_bytes = _regenerate_stamped_pdf_from_data(client, p)
                                # Save to temp and store path
                                import tempfile
                                stamped_path = tempfile.mktemp(suffix=".pdf", prefix=f"approved_{p['project_code']}_")
                                with open(stamped_path, 'wb') as f:
                                    f.write(pdf_bytes)
                                approve_project(p['id'], user['id'], stamped_path)
                                fname = f"{p.get('brand_name','')}-{p.get('project_code','')}-stamped.pdf"
                                st.session_state['just_approved'] = {
                                    'name': fname, 'path': stamped_path,
                                    'brand': p.get('brand_name',''),
                                    'code': p['project_code']
                                }
                                st.rerun()
                            except Exception as e:
                                st.error(f"生成失败: {e}")

                if st.button("❌ 驳回", key=f"reject_{p['id']}", use_container_width=True):
                    reject_project(p['id'], user['id'])
                    st.warning("已驳回")
                    st.rerun()

    # Show just-approved download banner
    if 'just_approved' in st.session_state and st.session_state['just_approved']:
        ja = st.session_state['just_approved']
        st.divider()
        st.success(f"✅ 审核通过！{ja['brand']} ({ja['code']}) 盖章 PDF 已生成")
        with open(ja['path'], "rb") as f:
            st.download_button(
                f"📥 下载盖章 PDF — {ja['name']}",
                f, file_name=ja['name'],
                key="dl_just_approved",
                use_container_width=True
            )
        if st.button("✅ 已下载，清除提示"):
            st.session_state['just_approved'] = None
            st.rerun()

    # Payment tracking: approved projects with cost info
    st.divider()
    st.subheader("💰 已通过项目 & 到账跟踪")
    all_projects = get_projects(limit=200)
    approved = [p for p in all_projects if p.get('status') == 'approved']
    if approved:
        import pandas as pd
        rows = []
        for p in approved:
            rows.append({
                '项目编号': p.get('project_code',''),
                '品牌': p.get('brand_name',''),
                '客户': p.get('client_short',''),
                '项目金额': f"{p.get('currency','USD')} {p.get('amount',0):,.0f}",
                '预估成本': f"{p.get('cost_currency','USD')} {p.get('estimated_cost',0):,.0f}",
                '成本构成': (p.get('cost_breakdown','') or '')[:60],
                '提交时间': (p.get('created_at','') or '')[:10],
                '已到账': '✅' if p.get('payment_received') else '⏳',
                '到账日期': str(p.get('received_date',''))[:10] if p.get('received_date') else '',
            })
        df = pd.DataFrame(rows)
        st.dataframe(df, use_container_width=True, hide_index=True)

        # Mark as received
        st.subheader("标记到账")
        not_received = [p for p in approved if not p.get('payment_received')]
        if not_received:
            col_a, col_b, col_c = st.columns([3, 1, 1])
            with col_a:
                proj_options = {f"{p['project_code']} - {p['brand_name']} ({p.get('currency','USD')} {p.get('amount',0):,.0f})": p for p in not_received}
                selected = st.selectbox("选择项目", list(proj_options.keys()), key="rcv_select")
            with col_b:
                rcv_date = st.date_input("到账日期", value=datetime.now(), key="rcv_date")
            with col_c:
                rcv_amount = st.number_input("到账金额", min_value=0.0, value=None, step=100.0, key="rcv_amount")
                if st.button("✅ 标记到账", type="primary", use_container_width=True):
                    p = proj_options[selected]
                    sb = get_connection()
                    sb.table("projects").update({
                        "payment_received": True,
                        "received_date": rcv_date.strftime("%Y-%m-%d"),
                        "received_amount": rcv_amount or p.get('amount', 0)
                    }).eq("id", p['id']).execute()
                    st.success(f"已标记 {p['brand_name']} 到账！")
                    st.rerun()

    # History of approved/rejected
    st.divider()
    st.subheader("审核历史")


# ============================================================
# Cash Receipt Page
# ============================================================
def page_receipt():
    st.title("🧾 开具收据")
    user = st.session_state.user
    if user['role'] not in ('finance', 'admin'):
        st.error("无权访问，仅财务角色可操作")
        return

    # Check if coming from workspace with a pre-selected project
    pre_sel_id = st.session_state.pop('receipt_project_id', None)
    pre_sel_proj = get_project_by_id(pre_sel_id) if pre_sel_id else None

    # Get approved projects to reference
    all_projects = get_projects(limit=200)
    approved = [p for p in all_projects if p.get('status') == 'approved']

    clients = get_clients()
    client_map = {c['id']: c for c in clients}

    mode = st.radio("选择方式", ["从已通过项目生成", "手动填写"], horizontal=True)

    if pre_sel_proj and mode == "从已通过项目生成":
        # Auto-select the project from workspace
        st.info(f"📌 已选择项目：**{pre_sel_proj.get('brand_name','')}** ({pre_sel_proj.get('project_code','')})")

    if pre_sel_proj:
        client = client_map.get(pre_sel_proj['client_id'], {})
        _receipt_form(client, pre_sel_proj)
    elif mode == "从已通过项目生成" and approved:
        proj_options = {f"{p['brand_name']} — {p['project_code']} ({p.get('currency','USD')} {p['amount']:,.0f})": p
                        for p in approved}
        selected_label = st.selectbox("选择已通过的项目", list(proj_options.keys()))
        selected_proj = proj_options[selected_label]
        client = client_map.get(selected_proj['client_id'], {})
        _receipt_form(client, selected_proj)
    elif mode == "从已通过项目生成":
        st.info("暂无已通过的项目，请先审核通过 Invoice 后再开收据")
    else:
        # Manual mode
        client_names = [c['short_name'] for c in clients]
        sel_client = st.selectbox("选择客户", client_names)
        client = {c['short_name']: c for c in clients}.get(sel_client, {})
        _receipt_form(client, None)


def _receipt_form(client, project):
    if not client:
        st.warning("请先选择客户")
        return

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("客户信息")
        if client:
            st.write(f"**{client.get('full_name', '')}**")
            st.caption(f"联系人: {client.get('contact', '')}")
            st.caption(f"地址: {client.get('address', '')}")

        if project:
            st.subheader("关联项目")
            st.write(f"**{project.get('brand_name', '')}** — {project.get('project_name', '')}")
            st.caption(f"编号: {project.get('project_code', '')}")
            st.caption(f"金额: {project.get('currency','USD')} {project.get('amount',0):,.2f}")
            default_project_name = project.get('project_name', '')
            default_project_code = project.get('project_code', '')
            default_amount = project['amount']
            default_currency = project.get('currency', 'USD')
            default_venue = project.get('venue', '')
            default_period = project.get('execution_period', '')
        else:
            st.subheader("项目信息（手动填写）")
            default_project_name = st.text_input("项目名称 *", placeholder="如 XXX品牌 7月UGC")
            default_project_code = st.text_input("项目编号", placeholder="如 WELL26070101")
            default_venue = st.text_input("地点", value="Bangkok")
            default_period = st.text_input("项目日期范围", value="Jul 2026")
            default_amount = 0.0
            default_currency = 'USD'

    with col2:
        st.subheader("收款信息")
        issuer = st.text_input("开具人", value="Mr. Terry.Su")
        payment_date = st.date_input("付款日期", value=datetime.now())
        gained_date = st.date_input("到款日期", value=datetime.now())
        payment_method = st.selectbox("付款方式", ["BANK", "CASH", "TRANSFER"])
        payment_amount = st.number_input("实收金额", min_value=0.0,
                                         value=float(default_amount) if default_amount > 0 else None, step=100.0)
        currency = st.selectbox("币种", ["USD", "RMB"],
                                index=0 if default_currency == "USD" else 1)

    st.divider()
    if st.button("🧾 生成收据并盖章", type="primary", use_container_width=True):
        if payment_amount <= 0:
            st.error("请输入收款金额")
            return

        with st.spinner("正在生成收据..."):
            receipt_data = {
                'client_short': client.get('short_name', ''),
                'brand_name': project.get('brand_name', 'N/A') if project else client.get('short_name', 'Manual'),
                'project_code': default_project_code,
                'project_name': default_project_name,
                'amount': default_amount,
                'project_date': default_period,
                'venue': default_venue,
                'currency': currency,
                'payment_amount': payment_amount,
                'payment_date': payment_date,
                'gained_date': gained_date,
                'payment_method': payment_method,
                'issuer_name': issuer,
            }

            # 1) Generate xlsx from correct template
            import tempfile, io
            try:
                xlsx_path = generate_cash_receipt(
                    {'full_name': client.get('full_name', ''), 'address': client.get('address', ''),
                     'contact': client.get('contact', ''), 'phone': client.get('phone', ''),
                     'email': client.get('email', '')},
                    receipt_data
                )
                with open(xlsx_path, 'rb') as fx:
                    st.session_state['receipt_xlsx'] = fx.read()
                st.session_state['receipt_xlsx_name'] = os.path.basename(xlsx_path)
            except Exception:
                st.session_state['receipt_xlsx'] = None

            # 2) Generate stamped PDF
            try:
                stamped_name = tempfile.mktemp(suffix=".pdf", prefix=f"receipt_{receipt_data['brand_name']}_")
                generate_receipt_pdf(
                    {'full_name': client['full_name'], 'address': client.get('address', ''),
                     'contact': client.get('contact', ''), 'phone': client.get('phone', ''),
                     'email': client.get('email', '')},
                    receipt_data,
                    stamped_name
                )
                st.success("✅ 收据已生成！")
                st.session_state['receipt_stamped'] = stamped_name
                st.session_state['receipt_brand'] = receipt_data.get('brand_name', 'receipt')
                st.rerun()
            except Exception as e:
                st.error(f"PDF 生成失败: {e}")

    # Show download if just generated
    if 'receipt_stamped' in st.session_state and st.session_state['receipt_stamped']:
        path = st.session_state['receipt_stamped']
        if os.path.exists(path):
            col_a, col_b = st.columns(2)
            with col_a:
                with open(path, "rb") as f:
                    fname = f"{st.session_state.get('receipt_brand','receipt')}-cash-receipt.pdf"
                    st.download_button("📥 下载盖章收据 PDF", f, file_name=fname, use_container_width=True)
            with col_b:
                if st.session_state.get('receipt_xlsx'):
                    st.download_button("📥 下载收据 Excel", st.session_state['receipt_xlsx'],
                                      file_name=st.session_state.get('receipt_xlsx_name', 'receipt.xlsx'),
                                      use_container_width=True,
                                      mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            if st.button("清除提示"):
                st.session_state['receipt_stamped'] = None
                st.session_state['receipt_xlsx'] = None
                st.rerun()


# ============================================================
# Cost Overview Page (finance/admin)
# ============================================================
def page_cost():
    st.title("📊 成本总览")
    user = st.session_state.user
    if user['role'] not in ('finance', 'admin'):
        st.error("无权访问")
        return

    all_projects = get_projects(limit=500)
    if not all_projects:
        st.info("暂无项目")
        return

    import pandas as pd
    rows = []
    for p in all_projects:
        rows.append({
            '项目编号': p.get('project_code',''),
            '品牌': p.get('brand_name',''),
            '客户': p.get('client_short',''),
            '状态': {'draft':'草稿','pending':'待审','approved':'通过','rejected':'驳回'}.get(p.get('status',''), p.get('status','')),
            '项目金额': f"{p.get('currency','USD')} {p.get('amount',0):,.0f}",
            '预估成本': f"{p.get('cost_currency','USD')} {p.get('estimated_cost',0):,.0f}",
            '预计到账': str(p.get('expected_payment_date',''))[:10] if p.get('expected_payment_date') else '-',
            '成本构成': (p.get('cost_breakdown','') or '')[:80],
            '提交时间': (p.get('created_at','') or '')[:10],
            '提交人': p.get('created_by_name',''),
            '已到账': '✅' if p.get('payment_received') else '⏳',
        })

    df = pd.DataFrame(rows)
    st.dataframe(df, use_container_width=True, hide_index=True)

    # Summary
    total_revenue = sum(p.get('amount',0) or 0 for p in all_projects)
    total_cost = sum(p.get('estimated_cost',0) or 0 for p in all_projects)
    approved = [p for p in all_projects if p.get('status')=='approved']
    approved_rev = sum(p.get('amount',0) or 0 for p in approved)
    approved_cost = sum(p.get('estimated_cost',0) or 0 for p in approved)

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("总项目数", len(all_projects))
    col2.metric("总收入 (全部)", f"${total_revenue:,.0f}")
    col3.metric("总成本 (全部)", f"${total_cost:,.0f}")
    col4.metric("预计到账项目", f"{sum(1 for p in all_projects if p.get('expected_payment_date'))}")

    if approved:
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("已通过项目", len(approved))
        col2.metric("已通过收入", f"${approved_rev:,.0f}")
        col3.metric("已通过成本", f"${approved_cost:,.0f}")
        col4.metric("已到账", f"{sum(1 for p in approved if p.get('payment_received'))}")


# ============================================================
# Project Proposal (立项审批)
# ============================================================
def page_proposal():
    st.title("📋 项目立项")
    user = st.session_state.user

    tab_submit, tab_review = st.tabs(["提交立项", "审批立项"]) if user['role'] in ('admin',) else st.tabs(["提交立项"])

    with tab_submit:
        st.subheader("提交新项目立项")
        clients = get_clients()
        client_names = [c['short_name'] for c in clients]
        sel = st.selectbox("客户简称", client_names, key="prop_client")
        c = {c['short_name']: c for c in clients}.get(sel, {})

        col1, col2 = st.columns(2)
        with col1:
            prop_name = st.text_input("项目名称 *", placeholder="如 XXX品牌7月UGC 150篇")
            prop_amount = st.number_input("预估金额", min_value=0.0, step=100.0)
            prop_currency = st.selectbox("币种", ["USD", "RMB"])
        with col2:
            prop_period = st.text_input("执行周期", placeholder="如 2026年8月")
            prop_posts = st.text_input("内容数量", placeholder="如 150篇")
            prop_platform = st.text_input("发布平台", value="小红书")

        prop_note = st.text_area("备注", placeholder="项目简要说明...")

        if st.button("📤 提交立项", type="primary", use_container_width=True):
            if not prop_name or not sel:
                st.error("请填写客户简称和项目名称")
            else:
                sb = get_connection()
                sb.table("projects").insert({
                    "project_code": f"PROP-{datetime.now().strftime('%Y%m%d%H%M%S')}",
                    "project_name": prop_name,
                    "client_id": c.get('id'),
                    "brand_name": sel,
                    "amount": prop_amount,
                    "currency": prop_currency,
                    "execution_period": prop_period,
                    "total_posts": prop_posts,
                    "platform": prop_platform,
                    "content_type": prop_note,
                    "status": "draft",
                    "proposal_status": "pending",
                    "created_by": user['id'],
                }).execute()
                st.success("✅ 立项申请已提交，等待审批！")
                st.rerun()

        # List my proposals
        st.divider()
        st.subheader("我的立项申请")
        all_proj = get_projects(limit=200)
        my_proposals = [p for p in all_proj if p.get('proposal_status') and p.get('created_by') == user['id']]
        if my_proposals:
            prop_map = {'pending': '⏳ 待审批', 'approved': '✅ 已通过', 'rejected': '❌ 已驳回'}
            for p in my_proposals[:20]:
                ps = prop_map.get(p.get('proposal_status'), p.get('proposal_status'))
                st.write(f"{ps} **{p.get('project_name','')}** — {p.get('client_short','')} — {p.get('created_at','')[:10]}")
        else:
            st.info("暂无立项申请")

    # Admin review tab
    if user['role'] in ('admin',):
        with tab_review:
            st.subheader("待审批立项")
            pending_props = [p for p in get_projects(limit=200) if p.get('proposal_status') == 'pending']
            if not pending_props:
                st.success("暂无待审批立项")
            else:
                for p in pending_props:
                    with st.container(border=True):
                        col1, col2 = st.columns([3, 1])
                        with col1:
                            st.write(f"**{p.get('project_name','')}**")
                            st.caption(f"客户: {p.get('client_short','')} | "
                                      f"金额: {p.get('currency','USD')} {p.get('amount',0):,.2f} | "
                                      f"周期: {p.get('execution_period','')} | "
                                      f"提交人: {p.get('created_by_name','')} | {p.get('created_at','')[:10]}")
                            if p.get('content_type'):
                                st.caption(f"备注: {p['content_type']}")
                        with col2:
                            if st.button("✅ 通过", key=f"app_p_{p['id']}", use_container_width=True):
                                sb = get_connection()
                                sb.table("projects").update({"proposal_status": "approved", "status": "draft"}).eq("id", p['id']).execute()
                                st.success("已通过！可生成文档")
                                st.rerun()
                            if st.button("❌ 驳回", key=f"rej_p_{p['id']}", use_container_width=True):
                                sb = get_connection()
                                sb.table("projects").update({"proposal_status": "rejected"}).eq("id", p['id']).execute()
                                st.warning("已驳回")
                                st.rerun()


# ============================================================
# Main routing
# ============================================================
if st.session_state.user is None:
    restored = _restore_session()
    if restored:
        st.session_state.user = restored
        st.session_state.page = "generate"

if st.session_state.user is None:
    page_login()
else:
    if not is_approved(st.session_state.user['id']) and st.session_state.user['role'] != 'admin':
        st.warning("⏳ 你的账号正在等待管理员审核，审核通过后即可使用。")
        st.button("退出", on_click=logout)
    else:
        render_sidebar()
        pages = {
            "workspace": page_workspace,
            "generate": page_generate,
            "clients": page_clients,
            "history": page_history,
            "overview": page_overview,
            "approval": page_approval,
            "receipt": page_receipt,
            "admin": page_admin,
        }
        page_fn = pages.get(st.session_state.page, page_generate)
        page_fn()
